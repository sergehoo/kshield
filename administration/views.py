"""KAYDAN SHIELD — Vues d'administration (back-office)."""
from __future__ import annotations

import logging

from django.contrib import messages as dj_messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse, reverse_lazy
from django.utils.functional import cached_property
from django.views import View
from django.views.generic import (
    CreateView, DetailView, FormView, TemplateView, UpdateView,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Helpers : filtres + recherche + pagination réutilisables
# ---------------------------------------------------------------------------
def apply_search(qs, q: str, fields: list[str]):
    """Applique une recherche `icontains` (OR sur chaque champ)."""
    if not q:
        return qs
    cond = Q()
    for f in fields:
        cond |= Q(**{f"{f}__icontains": q})
    return qs.filter(cond)


def apply_chip_filter(qs, chip: str, mapping: dict):
    """Applique un filtre nommé."""
    if not chip or chip == "all":
        return qs
    f = mapping.get(chip)
    if f is None:
        return qs
    if isinstance(f, Q):
        return qs.filter(f)
    if callable(f):
        return f(qs)
    return qs.filter(**f)


def paginate(qs, request, per_page: int = 25, page_param: str = "page"):
    """Retourne `(page_obj, page_qs)` à partir d'un queryset."""
    from django.core.paginator import EmptyPage, Paginator
    paginator = Paginator(qs, per_page)
    raw = request.GET.get(page_param) or "1"
    try:
        page_num = int(raw)
    except (TypeError, ValueError):
        page_num = 1
    try:
        page_obj = paginator.page(page_num)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages or 1)
    return page_obj, list(page_obj.object_list)


class BaseAdminView(TemplateView):
    """Mixin de base pour toutes les pages admin.

    Toutes les vues qui en héritent exigent une session authentifiée. Un sous-
    type peut déclarer `permission_required = "module.action"` pour exiger une
    permission RBAC en plus.
    """

    active_nav: str = ""
    page_title: str = "KAYDAN SHIELD"
    page_subtitle: str = ""
    breadcrumb: str = ""
    template_name: str = "administration/page.html"
    permission_required = None  # par défaut : juste login

    def dispatch(self, request, *args, **kwargs):
        # Login required — sauf si on autorise les anonymes explicitement
        if not request.user.is_authenticated:
            from django.shortcuts import redirect
            from urllib.parse import urlencode
            return redirect(f"/auth/login/?{urlencode({'next': request.get_full_path()})}")
        # RBAC optionnel
        codes = self.permission_required
        if codes:
            from accounts.rbac import user_has_any
            if isinstance(codes, str):
                codes = [codes]
            if not user_has_any(request.user, codes):
                dj_messages.error(request,
                    "Permission refusée — contactez votre administrateur.")
                from django.shortcuts import redirect
                return redirect("admin-dashboard")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update({
            "active_nav": self.active_nav,
            "page_title": self.page_title,
            "page_subtitle": self.page_subtitle,
            "breadcrumb": self.breadcrumb or self.page_title,
            "open_alerts_count": self._safe_count("antifraud", "FraudAlert", status="open"),
        })
        # Catalogue de permissions du user courant (utilisé par les templates
        # pour conditionner l'affichage des boutons)
        try:
            from accounts.rbac import user_permissions
            ctx["user_perms"] = user_permissions(self.request.user)
        except Exception:
            ctx["user_perms"] = set()
        return ctx

    @staticmethod
    def _safe_count(app_label: str, model_name: str, **filters) -> int:
        try:
            from django.apps import apps
            return apps.get_model(app_label, model_name).objects.filter(**filters).count()
        except Exception:
            return 0


# =============================================================================
# Pilotage
# =============================================================================
class DashboardView(BaseAdminView):
    template_name = "administration/dashboard.html"
    active_nav = "dashboard"
    page_title = "Tableau de bord"
    page_subtitle = "Suivi en direct des accès, des pointages et des alertes."
    breadcrumb = "Tableau de bord"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from collections import Counter
        from datetime import timedelta

        from django.db.models import Count
        from django.utils import timezone

        now = timezone.now()
        today = now.date()
        week_ago = now - timedelta(days=7)

        try:
            from access_control.models import AccessEvent
            from antifraud.models import FraudAlert
            from devices.models import Badge, BadgeScanEvent
            from employees.models import Employee
            from ouvriers.models import Worker
            from sites.models import Site
            from visitors.models import Visitor, VisitRequest

            # ── KPIs principaux ──────────────────────────────────────────
            ctx["kpi_employees_total"] = Employee.objects.count()
            ctx["kpi_employees_active"] = Employee.objects.filter(status="active").count()
            ctx["kpi_workers_total"] = Worker.objects.count()
            ctx["kpi_workers_active"] = Worker.objects.filter(status="active").count()
            ctx["kpi_visitors_total"] = Visitor.objects.count()
            ctx["kpi_visitors_today"] = VisitRequest.objects.filter(
                created_at__date=today,
            ).count()
            ctx["kpi_badges_active"] = Badge.objects.filter(
                status__in=("active", "assigned"),
            ).count()
            ctx["kpi_badges_total"] = Badge.objects.count()
            ctx["kpi_sites_active"] = Site.objects.filter(status="active").count()

            # ── Activité d'accès ─────────────────────────────────────────
            ctx["scans_today"] = AccessEvent.objects.filter(timestamp__date=today).count()
            ctx["scans_today_granted"] = AccessEvent.objects.filter(
                timestamp__date=today, decision="granted").count()
            ctx["scans_today_denied"] = AccessEvent.objects.filter(
                timestamp__date=today, decision="denied").count()
            ctx["scans_today_review"] = AccessEvent.objects.filter(
                timestamp__date=today, decision="review").count()
            ctx["scans_last_hour"] = AccessEvent.objects.filter(
                timestamp__gte=now - timedelta(hours=1)).count()
            yesterday_count = AccessEvent.objects.filter(
                timestamp__date=today - timedelta(days=1)).count() or 1
            ctx["scans_today_delta"] = round(
                (ctx["scans_today"] - yesterday_count) / yesterday_count * 100, 1)

            # ── Tendance 7 jours (mini-chart) ───────────────────────────
            scans_by_day = []
            max_count = 1
            for i in range(6, -1, -1):
                d = today - timedelta(days=i)
                c = AccessEvent.objects.filter(timestamp__date=d).count()
                scans_by_day.append({
                    "label": d.strftime("%a")[:3],
                    "date": d.strftime("%d/%m"),
                    "count": c,
                })
                max_count = max(max_count, c)
            for entry in scans_by_day:
                entry["height_pct"] = round(entry["count"] / max_count * 100, 1)
            ctx["scans_by_day"] = scans_by_day
            ctx["scans_7d_total"] = sum(e["count"] for e in scans_by_day)

            # ── Alertes anti-fraude ─────────────────────────────────────
            ctx["alerts_open"] = FraudAlert.objects.filter(
                status="open").count()
            ctx["alerts_critical"] = FraudAlert.objects.filter(
                status="open", severity__in=("critical", "high")).count()
            ctx["recent_alerts"] = list(
                FraudAlert.objects.select_related("rule", "site")
                .order_by("-raised_at")[:5]
            )

            # ── Top sites par activité (7 derniers jours) ───────────────
            site_counts = (AccessEvent.objects
                           .filter(timestamp__gte=week_ago)
                           .values("site_id", "site__name", "site__type")
                           .annotate(c=Count("id"))
                           .order_by("-c")[:5])
            ctx["top_sites"] = list(site_counts)

            # ── Présence en cours (uniques par holder dans la dernière h) ─
            recent_events = (AccessEvent.objects
                             .filter(timestamp__gte=now - timedelta(hours=8),
                                     decision="granted",
                                     holder_object_id__isnull=False)
                             .values("holder_kind", "holder_object_id"))
            unique_present = {(e["holder_kind"], e["holder_object_id"])
                              for e in recent_events}
            kind_distribution = Counter(k for k, _ in unique_present)
            ctx["present_total"] = len(unique_present)
            ctx["present_employees"] = kind_distribution.get("employee", 0)
            ctx["present_workers"] = kind_distribution.get("worker", 0)
            ctx["present_visitors"] = kind_distribution.get("visitor", 0)

            # ── Derniers événements ─────────────────────────────────────
            ctx["recent_events"] = _annotate_events_with_holders(list(
                AccessEvent.objects.select_related("site", "device")
                .order_by("-timestamp")[:8]
            ))

            # ── Visites prévues aujourd'hui ─────────────────────────────
            ctx["upcoming_visits"] = list(
                VisitRequest.objects.filter(
                    scheduled_at__date=today,
                    status__in=("pending", "approved"),
                ).select_related("visitor", "site", "host_employee")
                .order_by("scheduled_at")[:5]
            )

            # ── Badges par catégorie ────────────────────────────────────
            ctx["badges_visitor"] = Badge.objects.filter(category="visitor_qr").count()
            ctx["badges_employee"] = Badge.objects.filter(category="employee_rfid").count()
            ctx["badges_worker"] = Badge.objects.filter(category="worker_rfid").count()

        except Exception:
            import logging
            logging.getLogger(__name__).exception("DashboardView context failed")
            # Fallback pour ne pas planter le rendering
            for k in ("kpi_employees_total", "kpi_employees_active",
                      "kpi_workers_total", "kpi_workers_active",
                      "kpi_visitors_total", "kpi_visitors_today",
                      "kpi_badges_active", "kpi_badges_total", "kpi_sites_active",
                      "scans_today", "scans_today_granted", "scans_today_denied",
                      "scans_today_review", "scans_last_hour", "scans_today_delta",
                      "scans_7d_total", "alerts_open", "alerts_critical",
                      "present_total", "present_employees", "present_workers",
                      "present_visitors",
                      "badges_visitor", "badges_employee", "badges_worker"):
                ctx[k] = 0
            for k in ("scans_by_day", "recent_alerts", "top_sites",
                      "recent_events", "upcoming_visits"):
                ctx[k] = []
        return ctx


class AdminHomeView(DashboardView):
    pass


class RealtimeView(BaseAdminView):
    template_name = "administration/realtime.html"
    active_nav = "realtime"
    page_title = "Flux temps réel"
    page_subtitle = "Tous les scans badges, casques et caméras au fil de l'eau."
    breadcrumb = "Flux temps réel"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        try:
            from datetime import timedelta

            from django.utils import timezone

            from access_control.models import AccessEvent
            from sites.models import Site
            now = timezone.now()

            qs = _realtime_filtered_qs(self.request)
            ctx["filtered_count"] = qs.count()

            page_obj, page_qs = paginate(qs, self.request, per_page=50)
            ctx["page_obj"] = page_obj
            ctx["events"] = _annotate_events_with_holders(page_qs)

            ctx["events_last_min"] = AccessEvent.objects.filter(
                timestamp__gte=now - timedelta(minutes=1),
            ).count()
            ctx["events_last_5min"] = AccessEvent.objects.filter(
                timestamp__gte=now - timedelta(minutes=5),
            ).count()
            ctx["sites"] = Site.objects.order_by("name")
            ctx["denied_last_hour"] = AccessEvent.objects.filter(
                timestamp__gte=now - timedelta(hours=1),
                decision="denied",
            ).count()

            ctx["f_period"] = self.request.GET.get("period", "today")
            ctx["f_holder_kind"] = self.request.GET.get("holder_kind", "")
            ctx["f_site"] = self.request.GET.get("site", "")
            ctx["f_decision"] = self.request.GET.get("decision", "")
            ctx["f_individual"] = self.request.GET.get("individual", "")
            ctx["f_start"] = self.request.GET.get("start", "")
            ctx["f_end"] = self.request.GET.get("end", "")
        except Exception:
            ctx["events"] = []
            ctx["events_last_min"] = 0
            ctx["events_last_5min"] = 0
            ctx["denied_last_hour"] = 0
            ctx["sites"] = []
            ctx["filtered_count"] = 0
            ctx["page_obj"] = None
            ctx["f_period"] = "today"
            for k in ("f_holder_kind", "f_site", "f_decision", "f_individual",
                      "f_start", "f_end"):
                ctx[k] = ""
        return ctx


def _realtime_filtered_qs(request):
    """Construit le queryset AccessEvent filtré selon les params GET."""
    from datetime import datetime, timedelta

    from django.utils import timezone

    from access_control.models import AccessEvent

    now = timezone.now()
    period = request.GET.get("period", "today")
    qs = AccessEvent.objects.select_related("site", "device")

    if period == "today":
        qs = qs.filter(timestamp__date=now.date())
    elif period == "week":
        qs = qs.filter(timestamp__gte=now - timedelta(days=7))
    elif period == "month":
        qs = qs.filter(timestamp__gte=now - timedelta(days=30))
    elif period == "custom":
        start_s = request.GET.get("start")
        end_s = request.GET.get("end")
        if start_s:
            try:
                start = datetime.fromisoformat(start_s)
                qs = qs.filter(timestamp__gte=timezone.make_aware(start)
                               if timezone.is_naive(start) else start)
            except (ValueError, TypeError):
                pass
        if end_s:
            try:
                end = datetime.fromisoformat(end_s)
                end = timezone.make_aware(end) if timezone.is_naive(end) else end
                qs = qs.filter(timestamp__lte=end + timedelta(days=1))
            except (ValueError, TypeError):
                pass

    holder_kind = request.GET.get("holder_kind", "")
    if holder_kind in ("employee", "worker", "visitor"):
        qs = qs.filter(holder_kind=holder_kind)

    site_id = request.GET.get("site", "")
    if site_id and site_id.isdigit():
        qs = qs.filter(site_id=int(site_id))

    decision = request.GET.get("decision", "")
    if decision in ("granted", "denied", "review"):
        qs = qs.filter(decision=decision)

    individual = (request.GET.get("individual") or "").strip()
    if individual:
        qs = qs.filter(badge_uid__icontains=individual)

    return qs.order_by("-timestamp")


def _annotate_events_with_holders(events):
    """Charge les holders en batch et annote chaque event (évite N+1)."""
    by_kind = {}
    for e in events:
        if e.holder_kind and e.holder_object_id:
            by_kind.setdefault(e.holder_kind, set()).add(e.holder_object_id)

    holders_by_kind = {}
    if "employee" in by_kind:
        from employees.models import Employee
        holders_by_kind["employee"] = {
            o.id: o for o in Employee.objects.select_related("company")
            .filter(id__in=by_kind["employee"])
        }
    if "worker" in by_kind:
        from ouvriers.models import Worker
        holders_by_kind["worker"] = {
            o.id: o for o in Worker.objects.select_related("subcontractor", "trade")
            .filter(id__in=by_kind["worker"])
        }
    if "visitor" in by_kind:
        from visitors.models import Visitor
        holders_by_kind["visitor"] = {
            o.id: o for o in Visitor.objects.filter(id__in=by_kind["visitor"])
        }

    for e in events:
        e.holder_obj = None
        e.holder_label = ""
        if e.holder_kind and e.holder_object_id:
            h = holders_by_kind.get(e.holder_kind, {}).get(e.holder_object_id)
            if h:
                e.holder_obj = h
                if e.holder_kind == "visitor":
                    e.holder_label = f"{h.first_name} {h.last_name}"
                else:
                    e.holder_label = f"{h.first_name} {h.last_name} ({h.matricule})"
    return events


# =============================================================================
# Identités
# =============================================================================
class EmployeesView(BaseAdminView):
    template_name = "administration/employees.html"
    active_nav = "employees"
    page_title = "Employés"
    page_subtitle = "Annuaire RH des porteurs de badge NFC."
    breadcrumb = "Employés"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        q = (self.request.GET.get("q") or "").strip()
        chip = self.request.GET.get("f", "all")
        company_filter = self.request.GET.get("company", "")
        try:
            from accounts.scoping import scope_queryset_by_company
            from core.models import Company
            from devices.models import Badge
            from employees.models import Employee

            qs = Employee.objects.select_related("company", "department", "position")
            qs = apply_search(qs, q, ["first_name", "last_name", "matricule", "email"])
            # RBAC : restreint aux filiales du user (sauf super-admin)
            qs = scope_queryset_by_company(qs, self.request.user, "company")

            if company_filter and company_filter.isdigit():
                qs = qs.filter(company_id=int(company_filter))

            badged_ids = list(Badge.objects.filter(
                category="employee_rfid", holder_kind="employee",
                status__in=("active", "assigned"),
            ).values_list("holder_object_id", flat=True))

            qs = apply_chip_filter(qs, chip, {
                "all": None,
                "active": Q(status="active"),
                "on_leave": Q(status="on_leave"),
                "terminated": Q(status="terminated"),
                "field": Q(work_location__in=("field", "both")),
                "office": Q(work_location="office"),
                "with_badge": Q(id__in=badged_ids),
                "without_badge": ~Q(id__in=badged_ids) & Q(status="active"),
            }).order_by("last_name", "first_name")

            ctx["filtered_count"] = qs.count()
            page_obj, page_qs = paginate(qs, self.request, per_page=25)
            ctx["page_obj"] = page_obj
            ctx["employees"] = page_qs
            ctx["total"] = Employee.objects.count()
            ctx["with_badge_count"] = Employee.objects.filter(id__in=badged_ids).count()
            ctx["facial_count"] = Employee.objects.filter(
                face_profiles__is_active=True).distinct().count()
            ctx["sites_count"] = Employee.objects.exclude(
                authorized_sites__isnull=True).distinct().count()
            ctx["companies"] = Company.objects.filter(is_active=True).order_by("name")
            ctx["company_filter"] = company_filter
        except Exception:
            ctx["page_obj"] = None
            ctx["employees"] = []
            ctx["total"] = 0
            ctx["filtered_count"] = 0
            ctx["with_badge_count"] = 0
            ctx["facial_count"] = 0
            ctx["sites_count"] = 0
            ctx["companies"] = []
            ctx["company_filter"] = ""

        ctx["q"] = q
        ctx["active_chip"] = chip
        ctx["chips"] = [
            ("all", "Tous"), ("active", "Actifs"),
            ("on_leave", "En congé"), ("terminated", "Sortis"),
            ("field", "Chantier"), ("office", "Bureau"),
            ("with_badge", "Avec badge"), ("without_badge", "Sans badge"),
        ]
        return ctx


class WorkersView(BaseAdminView):
    template_name = "administration/workers.html"
    active_nav = "workers"
    page_title = "Ouvriers"
    page_subtitle = "Segment chantier — Badge UHF + casque connecté."
    breadcrumb = "Ouvriers"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        q = (self.request.GET.get("q") or "").strip()
        chip = self.request.GET.get("f", "all")
        subcontractor_filter = self.request.GET.get("subcontractor", "")
        tab = self.request.GET.get("tab", "workers")
        ctx["tab"] = tab if tab in ("workers", "certifications", "crews",
                                      "assignments", "subcontractors") else "workers"

        # Onglets secondaires
        try:
            from django.utils import timezone

            from ouvriers.models import (Crew, Subcontractor,
                                          WorkerAssignment, WorkerCertification)

            if ctx["tab"] == "certifications":
                cqs = (WorkerCertification.objects
                       .select_related("worker")
                       .order_by("-issued_at"))
                page_obj, page_qs = paginate(cqs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["certifications"] = page_qs
                ctx["expiring_soon"] = WorkerCertification.objects.filter(
                    valid_until__lte=timezone.now().date() + __import__("datetime").timedelta(days=30),
                    valid_until__gte=timezone.now().date(),
                ).count()
            elif ctx["tab"] == "crews":
                cqs = (Crew.objects.select_related("site", "foreman")
                       .prefetch_related("workers")
                       .order_by("-is_active", "site__name"))
                page_obj, page_qs = paginate(cqs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["crews"] = page_qs
            elif ctx["tab"] == "assignments":
                aqs = (WorkerAssignment.objects
                       .select_related("worker", "site", "crew")
                       .order_by("-is_active", "-started_at"))
                page_obj, page_qs = paginate(aqs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["assignments"] = page_qs
            elif ctx["tab"] == "subcontractors":
                sqs = Subcontractor.objects.order_by("-is_active", "name")
                page_obj, page_qs = paginate(sqs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["subcontractors_list"] = page_qs

            # Compteurs onglets
            ctx["count_certs"] = WorkerCertification.objects.count()
            ctx["count_crews"] = Crew.objects.filter(is_active=True).count()
            ctx["count_assignments"] = WorkerAssignment.objects.filter(is_active=True).count()
            ctx["count_subs"] = Subcontractor.objects.filter(is_active=True).count()
        except Exception:
            for k in ("certifications", "crews", "assignments", "subcontractors_list"):
                ctx[k] = []
            for k in ("count_certs", "count_crews", "count_assignments",
                      "count_subs", "expiring_soon"):
                ctx[k] = 0
        try:
            from accounts.scoping import get_user_company_ids
            from devices.models import Badge
            from ouvriers.models import Subcontractor, Worker

            qs = Worker.objects.select_related("trade", "subcontractor")
            qs = apply_search(qs, q, ["first_name", "last_name", "matricule",
                                       "phone", "id_document_number"])
            # RBAC : restreint via crews/assignments aux filiales du user
            company_ids = get_user_company_ids(self.request.user)
            if company_ids is not None:
                if not company_ids:
                    qs = qs.none()
                else:
                    qs = qs.filter(
                        Q(assignments__site__company_id__in=company_ids)
                        | Q(crews__site__company_id__in=company_ids)
                    ).distinct()

            if subcontractor_filter == "direct":
                qs = qs.filter(subcontractor__isnull=True)
            elif subcontractor_filter and subcontractor_filter.isdigit():
                qs = qs.filter(subcontractor_id=int(subcontractor_filter))

            badged_ids = set(Badge.objects.filter(
                category="worker_rfid", holder_kind="worker",
                status__in=("active", "assigned"),
            ).values_list("holder_object_id", flat=True))
            with_helmet_ids = set(Badge.objects.filter(
                category="worker_rfid", holder_kind="worker",
                status__in=("active", "assigned"), paired_helmet__isnull=False,
            ).values_list("holder_object_id", flat=True))

            qs = apply_chip_filter(qs, chip, {
                "all": None,
                "active": Q(status="active"),
                "suspended": Q(status="suspended"),
                "blacklisted": Q(status="blacklisted"),
                "with_badge": Q(id__in=list(badged_ids)),
                "without_badge": ~Q(id__in=list(badged_ids)) & Q(status="active"),
                "without_helmet": Q(id__in=list(badged_ids - with_helmet_ids)),
            }).order_by("last_name", "first_name")

            ctx["filtered_count"] = qs.count()
            page_obj, page_qs = paginate(qs, self.request, per_page=25)
            ctx["page_obj"] = page_obj
            workers = page_qs
            worker_ids = [w.id for w in workers]

            badges = (Badge.objects
                      .filter(category="worker_rfid", holder_kind="worker",
                              holder_object_id__in=worker_ids,
                              status__in=("active", "assigned", "suspended"))
                      .select_related("paired_helmet"))
            badges_by_worker = {b.holder_object_id: b for b in badges}
            for w in workers:
                w.active_badge = badges_by_worker.get(w.id)

            ctx["workers"] = workers
            ctx["total"] = Worker.objects.count()
            ctx["with_badge"] = len(badged_ids)
            ctx["with_helmet"] = len(with_helmet_ids)
            ctx["without_badge"] = Worker.objects.filter(status="active").exclude(id__in=list(badged_ids)).count()
            ctx["subcontractors"] = Subcontractor.objects.filter(is_active=True).order_by("name")
            ctx["subcontractor_filter"] = subcontractor_filter
        except Exception:
            ctx["page_obj"] = None
            ctx["workers"] = []
            ctx["total"] = 0
            ctx["filtered_count"] = 0
            ctx["with_badge"] = 0
            ctx["with_helmet"] = 0
            ctx["without_badge"] = 0
            ctx["subcontractors"] = []
            ctx["subcontractor_filter"] = ""

        ctx["q"] = q
        ctx["active_chip"] = chip
        ctx["chips"] = [
            ("all", "Tous"), ("active", "Actifs"),
            ("with_badge", "Avec badge"), ("without_badge", "Sans badge"),
            ("without_helmet", "Sans casque"),
            ("suspended", "Suspendus"), ("blacklisted", "Liste rouge"),
        ]
        return ctx


class VisitorsView(BaseAdminView):
    template_name = "administration/visitors.html"
    active_nav = "visitors"
    page_title = "Visiteurs"
    page_subtitle = "Pré-enregistrement, QR self-service, check-in et liste rouge."
    breadcrumb = "Visiteurs"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        q = (self.request.GET.get("q") or "").strip()
        chip = self.request.GET.get("f", "all")
        tab = self.request.GET.get("tab", "visitors")
        ctx["tab"] = tab if tab in ("visitors", "requests", "invitations",
                                      "passes", "watchlist", "purposes") else "visitors"
        try:
            from datetime import timedelta

            from django.utils import timezone

            from visitors.models import (Visitor, VisitorInvitation,
                                            VisitorPass, VisitPurpose,
                                            VisitRequest, Watchlist)

            # ── Onglets secondaires ─────────────────────────────────
            if ctx["tab"] == "requests":
                req_qs = (VisitRequest.objects
                          .select_related("visitor", "site", "host_employee", "purpose")
                          .order_by("-scheduled_at"))
                status = self.request.GET.get("status", "")
                if status:
                    req_qs = req_qs.filter(status=status)
                page_obj, page_qs = paginate(req_qs, self.request, per_page=20)
                ctx["page_obj"] = page_obj
                ctx["visit_requests"] = page_qs
                ctx["req_status"] = status
            elif ctx["tab"] == "invitations":
                inv_qs = (VisitorInvitation.objects
                          .select_related("visit_request", "visit_request__visitor")
                          .order_by("-created_at"))
                page_obj, page_qs = paginate(inv_qs, self.request, per_page=20)
                ctx["page_obj"] = page_obj
                ctx["invitations"] = page_qs
            elif ctx["tab"] == "passes":
                pass_qs = (VisitorPass.objects
                           .select_related("visit_request", "visit_request__visitor")
                           .order_by("-created_at"))
                page_obj, page_qs = paginate(pass_qs, self.request, per_page=20)
                ctx["page_obj"] = page_obj
                ctx["passes"] = page_qs
            elif ctx["tab"] == "watchlist":
                wl_qs = (Watchlist.objects.select_related("visitor", "site")
                         .order_by("-is_active", "-created_at"))
                page_obj, page_qs = paginate(wl_qs, self.request, per_page=20)
                ctx["page_obj"] = page_obj
                ctx["watchlist"] = page_qs
            elif ctx["tab"] == "purposes":
                p_qs = VisitPurpose.objects.order_by("-is_active", "code")
                page_obj, page_qs = paginate(p_qs, self.request, per_page=20)
                ctx["page_obj"] = page_obj
                ctx["purposes"] = page_qs

            from accounts.scoping import get_user_company_ids

            qs = Visitor.objects.all()
            qs = apply_search(qs, q, ["first_name", "last_name", "email",
                                       "phone", "id_number", "company"])

            # RBAC : visiteur visible si au moins une visite sur un site
            # de la filiale du user. Pour super-admin → tout.
            company_ids = get_user_company_ids(self.request.user)
            if company_ids is not None:
                if not company_ids:
                    qs = qs.none()
                else:
                    qs = qs.filter(
                        visit_requests__site__company_id__in=company_ids
                    ).distinct()

            today = timezone.now().date()
            # Limite aussi VisitRequest comptages aux sites du user
            vr_base = VisitRequest.objects.all()
            if company_ids is not None and company_ids:
                vr_base = vr_base.filter(site__company_id__in=company_ids)
            on_site_ids = list(vr_base.filter(
                status="checked_in",
            ).values_list("visitor_id", flat=True))
            recent_ids = list(vr_base.filter(
                created_at__gte=timezone.now() - timedelta(days=7),
            ).values_list("visitor_id", flat=True))

            qs = apply_chip_filter(qs, chip, {
                "all": None,
                "active": Q(pseudonymized_at__isnull=True),
                "anonymized": Q(pseudonymized_at__isnull=False),
                "on_site": Q(id__in=on_site_ids),
                "recent": Q(id__in=recent_ids),
            }).order_by("-created_at")

            ctx["filtered_count"] = qs.count()
            page_obj, page_qs = paginate(qs, self.request, per_page=25)
            ctx["page_obj"] = page_obj
            ctx["visitors"] = page_qs
            ctx["total_visitors"] = Visitor.objects.count()
            ctx["today_visits"] = VisitRequest.objects.filter(
                created_at__date=today,
            ).count()
            ctx["on_site_count"] = len(on_site_ids)
            ctx["watchlist_count"] = Visitor.objects.filter(pseudonymized_at__isnull=False).count()

            # Compteurs onglets pour la barre de tabs
            ctx["count_visitors"] = Visitor.objects.count()
            ctx["count_requests"] = VisitRequest.objects.count()
            ctx["count_pending"] = VisitRequest.objects.filter(status="pending").count()
            ctx["count_invitations"] = VisitorInvitation.objects.count()
            ctx["count_passes"] = VisitorPass.objects.filter(revoked_at__isnull=True).count()
            ctx["count_watchlist"] = Watchlist.objects.filter(is_active=True).count()
            ctx["count_purposes"] = VisitPurpose.objects.filter(is_active=True).count()
        except Exception:
            ctx["page_obj"] = None
            ctx["visitors"] = []
            ctx["filtered_count"] = 0
            ctx["total_visitors"] = 0
            ctx["today_visits"] = 0
            ctx["on_site_count"] = 0
            ctx["watchlist_count"] = 0
            for k in ("count_visitors", "count_requests", "count_pending",
                      "count_invitations", "count_passes", "count_watchlist",
                      "count_purposes"):
                ctx[k] = 0
            for k in ("visit_requests", "invitations", "passes", "watchlist", "purposes"):
                ctx[k] = []

        ctx["q"] = q
        ctx["active_chip"] = chip
        ctx["chips"] = [
            ("all", "Tous"), ("active", "Actifs"),
            ("on_site", "Sur site"), ("recent", "7 derniers jours"),
            ("anonymized", "Pseudonymisés"),
        ]
        return ctx


# =============================================================================
# Terrain
# =============================================================================
class SitesView(BaseAdminView):
    template_name = "administration/sites.html"
    active_nav = "sites"
    page_title = "Sites & zones"
    breadcrumb = "Sites & zones"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        try:
            from accounts.scoping import scope_queryset_by_company
            from sites.models import Site
            qs = Site.objects.all().order_by("name")
            # RBAC : restreint à la/les filiale(s) du user
            qs = scope_queryset_by_company(qs, self.request.user, "company")
            page_obj, page_qs = paginate(qs, self.request, per_page=25)
            ctx["page_obj"] = page_obj
            ctx["sites"] = page_qs
        except Exception:
            ctx["page_obj"] = None
            ctx["sites"] = []
        return ctx


class DevicesView(BaseAdminView):
    template_name = "administration/devices.html"
    active_nav = "devices"
    page_title = "Équipements"
    breadcrumb = "Équipements"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from datetime import timedelta

        from django.utils import timezone

        tab = self.request.GET.get("tab", "devices")
        ctx["tab"] = tab if tab in ("devices", "models", "heartbeats",
                                      "maintenance", "firmwares", "ota") else "devices"
        try:
            from devices.models import (Device, DeviceHeartbeat,
                                          DeviceMaintenance, DeviceModel,
                                          FirmwareVersion, OTAUpdate)

            if ctx["tab"] == "devices":
                qs = Device.objects.select_related("model", "site").order_by("serial_number")
                page_obj, page_qs = paginate(qs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["devices"] = page_qs
            elif ctx["tab"] == "models":
                mqs = DeviceModel.objects.order_by("brand", "model")
                page_obj, page_qs = paginate(mqs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["models"] = page_qs
            elif ctx["tab"] == "heartbeats":
                hqs = (DeviceHeartbeat.objects.select_related("device")
                       .order_by("-created_at"))
                page_obj, page_qs = paginate(hqs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["heartbeats"] = page_qs
            elif ctx["tab"] == "maintenance":
                mqs = (DeviceMaintenance.objects.select_related("device")
                       .order_by("-started_at"))
                page_obj, page_qs = paginate(mqs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["maintenances"] = page_qs
            elif ctx["tab"] == "firmwares":
                fqs = (FirmwareVersion.objects.select_related("device_model")
                       .order_by("-is_published", "-created_at"))
                page_obj, page_qs = paginate(fqs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["firmwares"] = page_qs
            elif ctx["tab"] == "ota":
                oqs = (OTAUpdate.objects.select_related("device", "firmware")
                       .order_by("-scheduled_for"))
                page_obj, page_qs = paginate(oqs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["ota_updates"] = page_qs

            # Compteurs onglets + devices hors-ligne (pas de heartbeat depuis 5 min)
            now = timezone.now()
            cutoff = now - timedelta(minutes=5)
            ctx["models_count"] = DeviceModel.objects.count()
            ctx["count_devices"] = Device.objects.count()
            ctx["count_heartbeats"] = DeviceHeartbeat.objects.count()
            ctx["count_maintenance"] = DeviceMaintenance.objects.filter(
                ended_at__isnull=True).count()
            ctx["count_firmwares"] = FirmwareVersion.objects.count()
            ctx["count_ota"] = OTAUpdate.objects.filter(
                status__in=("pending", "in_progress")).count()
            ctx["devices_offline"] = Device.objects.filter(
                last_heartbeat_at__isnull=False, last_heartbeat_at__lt=cutoff,
            ).count() + Device.objects.filter(last_heartbeat_at__isnull=True).count()
        except Exception:
            for k in ("devices", "models", "heartbeats", "maintenances",
                      "firmwares", "ota_updates"):
                ctx[k] = []
            for k in ("page_obj",):
                ctx[k] = None
            for k in ("models_count", "count_devices", "count_heartbeats",
                      "count_maintenance", "count_firmwares", "count_ota",
                      "devices_offline"):
                ctx[k] = 0
        return ctx


# ───────────────────────────────────────────────────────────────────────────
# Enrôlement en masse — badges (NFC/UHF/QR) et casques (UHF+BLE)
# ───────────────────────────────────────────────────────────────────────────
class BadgeEnrollmentView(BaseAdminView):
    """Hub d'enrôlement : saisie batch / CSV / scan live de badges et casques.

    Pas d'attribution ici — les badges créés sont en pool (status="available"),
    les casques en status="active" sans current_worker. L'attribution se fait
    ensuite via le workflow badge issue + pairing.
    """
    template_name = "administration/badge_enrollment.html"
    active_nav = "badges"
    page_title = "Enrôler badges & tags"
    page_subtitle = (
        "Pré-enregistrez vos cartes NFC, tags UHF et beacons BLE avant de les "
        "attribuer aux employés / ouvriers."
    )
    breadcrumb = "Badges · Enrôlement"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        try:
            from devices.models import Badge, Device, Helmet
            ctx["count_pool_badges"] = Badge.objects.filter(status="available").count()
            ctx["count_unassigned_helmets"] = Helmet.objects.filter(
                current_worker__isnull=True, status="active",
            ).count()
            # Lecteurs RFID/NFC/ZK actifs — utilisable pour la capture live
            ctx["readers"] = list(
                Device.objects
                .select_related("model")
                .filter(model__type__in=[
                    "reader_uhf_fixed", "reader_uhf_mobile",
                    "reader_nfc_fixed", "reader_nfc_mobile",
                    "portique",
                ], status="active")
                .order_by("model__brand", "serial_number")[:50]
            )
        except Exception:
            ctx["count_pool_badges"] = 0
            ctx["count_unassigned_helmets"] = 0
            ctx["readers"] = []
        return ctx


# ───────────────────────────────────────────────────────────────────────────
# Équipements — picker technologie + assistant création lecteur
# ───────────────────────────────────────────────────────────────────────────
class DeviceReaderPickerView(BaseAdminView):
    """Étape 1 : choix de la technologie du lecteur à enregistrer."""
    template_name = "administration/device_reader_picker.html"
    active_nav = "devices"
    page_title = "Nouvel équipement"
    page_subtitle = "Quel type de lecteur souhaitez-vous ajouter ?"
    breadcrumb = "Équipements · Nouveau"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from administration.forms import READER_KIND_META, READER_KIND_TYPES
        from devices.models import DeviceModel

        # Compteur de DeviceModel par techno → on peut prévenir "0 modèle dispo"
        counts = {}
        for kind, types in READER_KIND_TYPES.items():
            counts[kind] = DeviceModel.objects.filter(
                type__in=types, is_active=True,
            ).count()
        ctx["reader_kinds"] = [
            {"key": k, **meta, "model_count": counts.get(k, 0)}
            for k, meta in READER_KIND_META.items()
        ]
        return ctx


class DeviceReaderCreateView(BaseAdminView):
    """Étape 2 : formulaire de création pré-filtré sur la technologie choisie."""
    template_name = "administration/device_reader_form.html"
    active_nav = "devices"
    breadcrumb = "Équipements · Nouveau lecteur"

    # On override pour gérer GET (form vide) ET POST (validation) dans une seule View
    def dispatch(self, request, *args, **kwargs):
        # garde le contrôle login/RBAC du parent
        resp = super().dispatch(request, *args, **kwargs)
        return resp

    def _kind(self):
        from administration.forms import READER_KIND_META
        kind = (self.kwargs.get("kind") or "").lower()
        if kind not in READER_KIND_META:
            return None
        return kind

    def get(self, request, *args, **kwargs):
        from django.shortcuts import redirect
        kind = self._kind()
        if not kind:
            dj_messages.error(request, "Technologie de lecteur inconnue.")
            return redirect("admin-device-reader-picker")
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        from django.shortcuts import redirect

        from administration.forms import DeviceForm
        kind = self._kind()
        if not kind:
            dj_messages.error(request, "Technologie de lecteur inconnue.")
            return redirect("admin-device-reader-picker")

        form = DeviceForm(request.POST, reader_kind=kind)
        if form.is_valid():
            obj = form.save(commit=False)
            # Tenant auto (cf. InjectKaydanTenantMixin)
            try:
                from core.services import get_kaydan_tenant
                obj.tenant = get_kaydan_tenant()
            except Exception:
                logger.exception("Impossible d'attacher le tenant Kaydan au device")
            obj.save()
            dj_messages.success(
                request,
                f"Lecteur {obj.serial_number} enregistré avec succès.",
            )
            return redirect("admin-device-detail", pk=obj.pk)

        # Si validation KO → re-render avec erreurs
        ctx = self.get_context_data(form=form)
        return self.render_to_response(ctx)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from administration.forms import DeviceForm, READER_KIND_META

        kind = self._kind()
        meta = READER_KIND_META[kind]
        ctx["reader_kind"] = kind
        ctx["reader_meta"] = meta
        ctx["page_title"] = f"Nouveau {meta['label'].lower()}"
        ctx["page_subtitle"] = meta["hint"]

        if "form" not in ctx:
            # Pré-remplit depuis les query params (import depuis le scan réseau)
            initial = {}
            for key in ("serial_number", "ip_address", "mac_address",
                        "firmware_version"):
                v = self.request.GET.get(key)
                if v:
                    initial[key] = v
            ctx["form"] = DeviceForm(reader_kind=kind, initial=initial or None)
            ctx["imported_from_scan"] = bool(initial)
        return ctx


# ───────────────────────────────────────────────────────────────────────────
# Caméras IP — liste/config + vue live multi-flux
# ───────────────────────────────────────────────────────────────────────────
class CamerasView(BaseAdminView):
    """Liste + statut + actions de configuration des caméras IP."""
    template_name = "administration/cameras.html"
    active_nav = "cameras"
    page_title = "Caméras IP"
    page_subtitle = "Streaming temps réel + pipeline IA (visage, mouvement, recording)."
    breadcrumb = "Caméras IP"
    permission_required = "devices.view"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        try:
            from devices.models import Camera
            qs = Camera.objects.select_related("site", "zone").order_by(
                "-is_active", "name",
            )
            site_filter = self.request.GET.get("site") or ""
            status_filter = self.request.GET.get("status") or ""
            search = self.request.GET.get("q") or ""
            if site_filter:
                qs = qs.filter(site_id=site_filter)
            if status_filter:
                qs = qs.filter(status=status_filter)
            if search:
                qs = qs.filter(
                    Q(name__icontains=search)
                    | Q(location_label__icontains=search)
                    | Q(rtsp_url__icontains=search),
                )
            page_obj, page_qs = paginate(qs, self.request, per_page=24)
            ctx["page_obj"] = page_obj
            ctx["cameras"] = page_qs
            ctx["count_total"] = Camera.objects.count()
            ctx["count_online"] = Camera.objects.filter(status="online", is_active=True).count()
            ctx["count_offline"] = Camera.objects.filter(status="offline", is_active=True).count()
            ctx["count_error"] = Camera.objects.filter(status="error").count()
            ctx["count_face_ai"] = Camera.objects.filter(enable_face_recognition=True).count()
            from sites.models import Site
            ctx["sites_for_filter"] = Site.objects.order_by("name")
            ctx["site_filter"] = site_filter
            ctx["status_filter"] = status_filter
            ctx["search_q"] = search
        except Exception:
            logger.warning("CamerasView ctx", exc_info=True)
            ctx["cameras"] = []
            ctx["page_obj"] = None
            for k in ("count_total", "count_online", "count_offline",
                       "count_error", "count_face_ai"):
                ctx[k] = 0
            ctx["sites_for_filter"] = []
        return ctx


class FacePresenceView(BaseAdminView):
    """Liste des FaceSightingEvent + FaceCheckinConfirmation du jour."""
    template_name = "administration/face_presence.html"
    active_nav = "cameras"
    page_title = "Présence par reconnaissance faciale"
    page_subtitle = "Confirmation visage ↔ badge RFID · vue temps réel"
    breadcrumb = "Caméras · Présence face"
    permission_required = "attendance.view"

    def get_context_data(self, **kwargs):
        from datetime import timedelta
        ctx = super().get_context_data(**kwargs)
        try:
            from accounts.scoping import scope_queryset_by_company
            from attendance.models import (FaceCheckinConfirmation,
                                              FaceSightingEvent)

            now = timezone.now()
            today = timezone.localdate()
            day_param = self.request.GET.get("date") or ""
            try:
                if day_param:
                    today = datetime.strptime(day_param, "%Y-%m-%d").date()
            except ValueError:
                pass

            # Sightings du jour, scope filiale via camera.site.company
            sightings = FaceSightingEvent.objects.select_related(
                "camera", "site", "employee", "employee__company",
            ).filter(timestamp__date=today).order_by("-timestamp")
            sightings = scope_queryset_by_company(
                sightings, self.request.user, "camera__site__company")
            ctx["sightings"] = sightings[:200]

            # Confirmations du jour
            confs = FaceCheckinConfirmation.objects.select_related(
                "employee", "employee__company", "punch", "sighting",
            ).filter(date=today).order_by("-created_at")
            confs = scope_queryset_by_company(confs, self.request.user, "employee__company")
            ctx["confirmations"] = list(confs)

            # KPIs
            ctx["total_sightings"] = sightings.count()
            ctx["total_matched"] = sightings.filter(matched=True).count()
            ctx["total_unmatched"] = sightings.filter(matched=False).count()
            ctx["count_confirmed"] = sum(1 for c in confs if c.status == "confirmed")
            ctx["count_face_only"] = sum(1 for c in confs if c.status == "face_only")
            ctx["count_badge_only"] = sum(1 for c in confs if c.status == "badge_only")
            ctx["date"] = today.isoformat()
            ctx["date_label"] = today.strftime("%d/%m/%Y")
        except Exception:
            logger.warning("FacePresenceView ctx", exc_info=True)
            ctx["sightings"] = []
            ctx["confirmations"] = []
            for k in ("total_sightings", "total_matched", "total_unmatched",
                       "count_confirmed", "count_face_only", "count_badge_only"):
                ctx[k] = 0
        return ctx


class CamerasLiveView(BaseAdminView):
    """Vue plein écran multi-flux MJPEG (grille 1×1 → 4×4)."""
    template_name = "administration/cameras_live.html"
    active_nav = "cameras"
    page_title = "Live multi-caméras"
    breadcrumb = "Caméras · Live"
    permission_required = "devices.view"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        try:
            from devices.models import Camera
            qs = Camera.objects.filter(is_active=True).select_related("site")
            # Permet de pré-sélectionner certaines caméras via ?ids=1,2,3
            sel = self.request.GET.get("ids", "")
            if sel:
                ids = [int(x) for x in sel.split(",") if x.strip().isdigit()]
                if ids:
                    qs = qs.filter(pk__in=ids)
            ctx["cameras"] = list(qs.order_by("site__name", "name"))
            ctx["initial_grid"] = self.request.GET.get("grid", "2")
        except Exception:
            logger.warning("CamerasLiveView ctx", exc_info=True)
            ctx["cameras"] = []
            ctx["initial_grid"] = "2"
        return ctx


class BadgesView(BaseAdminView):
    template_name = "administration/badges.html"
    active_nav = "badges"
    page_title = "Badges & casques"
    page_subtitle = "Trois workflows : QR visiteur · RFID employé · RFID ouvrier (couplé casque)."
    breadcrumb = "Badges & casques"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from datetime import timedelta

        from django.db.models import Count
        from django.utils import timezone

        q = (self.request.GET.get("q") or "").strip()
        chip = self.request.GET.get("f", "all")

        def _apply_badge_filter(qs):
            qs = apply_search(qs, q, ["uid", "qr_payload"])
            return apply_chip_filter(qs, chip, {
                "all": None,
                "active": Q(status="active"),
                "available": Q(status="available"),
                "assigned": Q(status="assigned"),
                "suspended": Q(status="suspended"),
                "expired": Q(status="expired"),
                "lost": Q(status="lost"),
                "revoked": Q(status="revoked"),
                "disabled": Q(status="disabled"),
            })

        try:
            from devices.models import Badge, BadgeScanEvent, Helmet
            now = timezone.now()
            today = now.date()

            tab = self.request.GET.get("tab", "visitor")

            visitor_qs = _apply_badge_filter(
                Badge.objects.filter(category="visitor_qr")
            ).order_by("status", "-issued_at")
            employee_qs = _apply_badge_filter(
                Badge.objects.filter(category="employee_rfid")
            ).select_related("paired_helmet").order_by("-issued_at")
            worker_qs = _apply_badge_filter(
                Badge.objects.filter(category="worker_rfid")
            ).select_related("paired_helmet").order_by("-issued_at")

            ctx["tab"] = tab if tab in ("visitor", "employee", "worker", "helmet") else "visitor"
            target_qs = {
                "visitor": visitor_qs,
                "employee": employee_qs,
                "worker": worker_qs,
            }.get(ctx["tab"], visitor_qs)
            page_obj, page_qs = paginate(target_qs, self.request, per_page=20)
            ctx["page_obj"] = page_obj

            ctx["visitor_badges"] = page_qs if ctx["tab"] == "visitor" else list(visitor_qs[:50])
            ctx["employee_badges"] = page_qs if ctx["tab"] == "employee" else list(employee_qs[:50])
            ctx["worker_badges"] = page_qs if ctx["tab"] == "worker" else list(worker_qs[:50])

            # Onglet Casques : pagination dédiée
            helmet_qs = (Helmet.objects
                         .select_related("current_worker")
                         .annotate(active_pairings=Count(
                             "paired_badges", filter=Q(paired_badges__status__in=("active", "assigned"))))
                         .order_by("status", "serial_number"))
            if ctx["tab"] == "helmet":
                page_obj, page_qs = paginate(helmet_qs, self.request, per_page=20)
                ctx["page_obj"] = page_obj
                ctx["helmets"] = page_qs
            else:
                ctx["helmets"] = list(helmet_qs[:50])
            ctx["helmet_total"] = Helmet.objects.count()
            ctx["helmet_active"] = Helmet.objects.filter(status="active").count()
            ctx["visitor_pool_count"] = Badge.objects.filter(category="visitor_qr", status="available").count()
            ctx["visitor_assigned_count"] = Badge.objects.filter(category="visitor_qr", status="active").count()
            ctx["employee_total"] = Badge.objects.filter(category="employee_rfid").count()
            ctx["worker_total"] = Badge.objects.filter(category="worker_rfid").count()

            status_counts = dict(
                Badge.objects.values_list("status").annotate(c=Count("id")).values_list("status", "c")
            )
            ctx["stats_total"] = sum(status_counts.values())
            for k in ("active", "assigned", "available", "suspended", "expired",
                      "lost", "revoked", "disabled"):
                ctx[f"stats_{k}"] = status_counts.get(k, 0)

            week_ago = now - timedelta(days=7)
            ctx["scans_today"] = BadgeScanEvent.objects.filter(timestamp__date=today).count()
            ctx["scans_today_granted"] = BadgeScanEvent.objects.filter(
                timestamp__date=today, decision="granted").count()
            ctx["scans_today_denied"] = BadgeScanEvent.objects.filter(
                timestamp__date=today, decision="denied").count()
            ctx["scans_7d"] = BadgeScanEvent.objects.filter(timestamp__gte=week_ago).count()

            scans_by_day = []
            for i in range(6, -1, -1):
                d = today - timedelta(days=i)
                cnt = BadgeScanEvent.objects.filter(timestamp__date=d).count()
                scans_by_day.append({"date": d.strftime("%d/%m"), "count": cnt})
            ctx["scans_by_day"] = scans_by_day

            from employees.models import Employee
            from ouvriers.models import Worker

            employees_with_badge = Badge.objects.filter(
                category="employee_rfid", status__in=("active", "assigned"),
                holder_kind="employee",
            ).values_list("holder_object_id", flat=True)
            ctx["unbadged_employees"] = Employee.objects.filter(
                status="active",
            ).exclude(id__in=employees_with_badge).order_by("last_name")[:200]

            workers_with_badge = Badge.objects.filter(
                category="worker_rfid", status__in=("active", "assigned"),
                holder_kind="worker",
            ).values_list("holder_object_id", flat=True)
            ctx["unbadged_workers"] = Worker.objects.filter(
                status="active",
            ).exclude(id__in=workers_with_badge).order_by("last_name")[:200]

            ctx["available_helmets"] = Helmet.objects.filter(
                status="active",
            ).exclude(
                paired_badges__status__in=("active", "assigned"),
            ).distinct().order_by("serial_number")[:100]

            # Badges en pool (disponibles, non attribués) — pour le workflow
            # "Attribuer un badge enrôlé à un employé/ouvrier"
            ctx["pool_employee_badges"] = Badge.objects.filter(
                status="available",
                category__in=("employee_rfid", "worker_rfid"),
                holder_object_id__isnull=True,
            ).order_by("uid")[:200]
            ctx["pool_employee_count"] = ctx["pool_employee_badges"].count() if hasattr(
                ctx["pool_employee_badges"], "count") else len(ctx["pool_employee_badges"])

        except Exception:
            for k in ("visitor_badges", "employee_badges", "worker_badges",
                      "helmets", "unbadged_employees", "unbadged_workers",
                      "available_helmets", "scans_by_day"):
                ctx[k] = []
            for k in ("visitor_pool_count", "visitor_assigned_count",
                      "employee_total", "worker_total",
                      "stats_total", "stats_active", "stats_assigned",
                      "stats_available", "stats_suspended", "stats_expired",
                      "stats_lost", "stats_revoked", "stats_disabled",
                      "scans_today", "scans_today_granted",
                      "scans_today_denied", "scans_7d"):
                ctx[k] = 0
            ctx["tab"] = "visitor"
            ctx["page_obj"] = None

        ctx["q"] = q
        ctx["active_chip"] = chip
        ctx["badge_chips"] = [
            ("all", "Tous"), ("active", "Actifs"),
            ("available", "Disponibles"), ("assigned", "Attribués"),
            ("suspended", "Suspendus"), ("expired", "Expirés"),
            ("lost", "Perdus"), ("revoked", "Révoqués"),
            ("disabled", "Désactivés"),
        ]
        return ctx


class BadgeScanView(BaseAdminView):
    template_name = "administration/badge_scan.html"
    active_nav = "badges"
    page_title = "Scan badge en temps réel"
    breadcrumb = "Badges · Scanner"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        try:
            from datetime import timedelta

            from django.utils import timezone

            from devices.models import BadgeScanEvent
            since = timezone.now() - timedelta(hours=2)
            ctx["recent_scans"] = list(
                BadgeScanEvent.objects.filter(timestamp__gte=since)
                .select_related("badge", "site")
                .order_by("-timestamp")[:30]
            )
        except Exception:
            ctx["recent_scans"] = []
        return ctx


class GatewaysView(BaseAdminView):
    template_name = "administration/gateways.html"
    active_nav = "gateways"
    page_title = "Gateways locales"
    breadcrumb = "Gateways"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        try:
            from core.models import SiteGateway
            qs = SiteGateway.objects.select_related("site").order_by("-status", "site__name")
            page_obj, page_qs = paginate(qs, self.request, per_page=25)
            ctx["page_obj"] = page_obj
            ctx["gateways"] = page_qs
            ctx["total"] = SiteGateway.objects.count()
            ctx["online"] = SiteGateway.objects.filter(status="active").count()
            ctx["offline"] = SiteGateway.objects.filter(status="offline").count()
        except Exception:
            ctx["gateways"] = []
            ctx["page_obj"] = None
            ctx["total"] = 0
            ctx["online"] = 0
            ctx["offline"] = 0
        return ctx


# =============================================================================
# Pages secondaires (placeholders simples — étendables plus tard)
# =============================================================================
class AttendanceView(BaseAdminView):
    template_name = "administration/attendance.html"
    active_nav = "attendance"
    page_title = "Pointage & présence"
    page_subtitle = "Punches, journées de travail et demandes de congés."
    breadcrumb = "Pointage"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from datetime import timedelta

        from django.utils import timezone
        try:
            from attendance.models import (AttendanceDay, LeaveRequest,
                                            OvertimeRule, Punch)
            now = timezone.now()
            today = now.date()
            week_ago = today - timedelta(days=7)

            tab = self.request.GET.get("tab", "punches")
            ctx["tab"] = tab if tab in ("punches", "days", "leaves", "rules",
                                          "corrections", "rosters", "overtime",
                                          "sheet") else "punches"

            # ───── Tab Feuille de pointage : agrégation des AccessEvent ─────
            if ctx["tab"] == "sheet":
                ctx.update(self._build_sheet_context())

            if ctx["tab"] == "punches":
                qs = (Punch.objects.select_related("site")
                      .order_by("-timestamp"))
                page_obj, page_qs = paginate(qs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["punches"] = page_qs
            elif ctx["tab"] == "days":
                qs = (AttendanceDay.objects.select_related("site")
                      .order_by("-date"))
                page_obj, page_qs = paginate(qs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["days"] = page_qs
            elif ctx["tab"] == "leaves":
                qs = (LeaveRequest.objects.select_related("employee", "worker", "approved_by")
                      .order_by("-start_date"))
                page_obj, page_qs = paginate(qs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["leaves"] = page_qs
            elif ctx["tab"] == "rules":
                qs = (OvertimeRule.objects.select_related("company")
                      .order_by("-is_active", "company__name"))
                page_obj, page_qs = paginate(qs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["rules"] = page_qs
            elif ctx["tab"] == "corrections":
                from attendance.models import AttendanceCorrection
                cqs = (AttendanceCorrection.objects
                       .select_related("attendance_day", "performed_by")
                       .order_by("-created_at"))
                page_obj, page_qs = paginate(cqs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["corrections"] = page_qs
            elif ctx["tab"] == "rosters":
                from attendance.models import Roster
                rqs = Roster.objects.select_related("site").order_by("-date")
                page_obj, page_qs = paginate(rqs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["rosters"] = page_qs
            elif ctx["tab"] == "overtime":
                from attendance.models import OvertimeCalculation
                oqs = (OvertimeCalculation.objects
                       .select_related("employee", "worker")
                       .order_by("-week_start"))
                page_obj, page_qs = paginate(oqs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["overtime_calcs"] = page_qs

            # Compte pour le badge sur l'onglet "sheet"
            from access_control.models import AccessEvent
            ctx["events_today_count"] = AccessEvent.objects.filter(
                timestamp__date=today,
            ).count()

            ctx["punches_today"] = Punch.objects.filter(timestamp__date=today).count()
            ctx["punches_week"] = Punch.objects.filter(timestamp__date__gte=week_ago).count()
            ctx["days_today"] = AttendanceDay.objects.filter(date=today).count()
            ctx["leaves_pending"] = LeaveRequest.objects.filter(status="pending").count()
            ctx["rules_active"] = OvertimeRule.objects.filter(is_active=True).count()
        except Exception:
            for k in ("punches", "days", "leaves", "rules"):
                ctx[k] = []
            for k in ("punches_today", "punches_week", "days_today",
                      "leaves_pending", "rules_active"):
                ctx[k] = 0
            ctx["page_obj"] = None
            ctx["tab"] = "punches"
        return ctx

    def _build_sheet_context(self):
        """Agrège les AccessEvent par employé/ouvrier et par jour.

        Filtres GET acceptés :
          - period : "day" (default today), "week", "month", "custom"
          - start, end : dates ISO si period=custom
          - holder_kind : "" / "employee" / "worker"
          - site : id
          - q : recherche par nom/matricule
        """
        from datetime import date as _date, datetime, timedelta

        from django.contrib.contenttypes.models import ContentType
        from django.db.models import Max, Min
        from django.utils import timezone

        from access_control.models import AccessEvent
        from employees.models import Employee
        from ouvriers.models import Worker
        from sites.models import Site

        req = self.request
        period = req.GET.get("period", "day")
        site_id = req.GET.get("site") or ""
        holder_kind = req.GET.get("holder_kind") or ""
        q = (req.GET.get("q") or "").strip()

        now = timezone.now()
        today = now.date()

        # Période
        if period == "week":
            start_date = today - timedelta(days=6)
            end_date = today
        elif period == "month":
            start_date = today - timedelta(days=29)
            end_date = today
        elif period == "custom":
            try:
                start_date = datetime.fromisoformat(req.GET.get("start", "")).date()
            except Exception:
                start_date = today
            try:
                end_date = datetime.fromisoformat(req.GET.get("end", "")).date()
            except Exception:
                end_date = today
            if end_date < start_date:
                start_date, end_date = end_date, start_date
        else:    # day
            start_date = today
            end_date = today

        # Query base AccessEvent
        qs = AccessEvent.objects.filter(
            timestamp__date__gte=start_date,
            timestamp__date__lte=end_date,
        )
        if site_id:
            qs = qs.filter(site_id=site_id)
        if holder_kind:
            qs = qs.filter(holder_kind=holder_kind)

        # Agrégation : pour chaque (holder, date) → first_in, last_out
        # On utilise des values() + GROUP BY puis on enrichit
        from django.db.models.functions import TruncDate
        agg = (qs
               .exclude(holder_object_id__isnull=True)
               .annotate(day=TruncDate("timestamp"))
               .values("holder_kind", "holder_object_id", "day", "site")
               .annotate(first_in=Min("timestamp"), last_out=Max("timestamp"))
               .order_by("-day", "holder_kind", "holder_object_id"))

        # Résolve les holders (Employee/Worker) en bulk
        emp_ids = {r["holder_object_id"] for r in agg if r["holder_kind"] == "employee"}
        worker_ids = {r["holder_object_id"] for r in agg if r["holder_kind"] == "worker"}
        emp_map = {e.pk: e for e in Employee.objects.filter(pk__in=emp_ids)}
        worker_map = {w.pk: w for w in Worker.objects.filter(pk__in=worker_ids)}

        rows = []
        for r in agg:
            holder = (emp_map if r["holder_kind"] == "employee" else worker_map).get(
                r["holder_object_id"])
            if not holder:
                continue
            full_name = f"{holder.first_name} {holder.last_name}".strip()
            if q and q.lower() not in (
                holder.matricule.lower() + " " + full_name.lower()
            ):
                continue
            first = r["first_in"]
            last = r["last_out"]
            duration_min = 0
            if first and last and last > first:
                duration_min = int((last - first).total_seconds() // 60)
            rows.append({
                "date": r["day"],
                "holder_kind": r["holder_kind"],
                "holder": holder,
                "matricule": holder.matricule,
                "name": full_name,
                "first_in": first,
                "last_out": last,
                "duration_min": duration_min,
                "duration_hms": self._fmt_duration(duration_min),
                "site_id": r["site"],
            })

        # Stats
        total_persons = len({(r["holder_kind"], r["holder"].pk) for r in rows})
        total_days = len({r["date"] for r in rows})
        total_minutes = sum(r["duration_min"] for r in rows)

        return {
            "sheet_rows": rows,
            "sheet_period": period,
            "sheet_start": start_date.isoformat(),
            "sheet_end": end_date.isoformat(),
            "sheet_site_id": site_id,
            "sheet_holder_kind": holder_kind,
            "sheet_q": q,
            "sheet_sites": list(Site.objects.order_by("name").only("id", "name")),
            "sheet_total_persons": total_persons,
            "sheet_total_days": total_days,
            "sheet_total_hours": round(total_minutes / 60.0, 1),
        }

    @staticmethod
    def _fmt_duration(minutes: int) -> str:
        if not minutes:
            return "—"
        h, m = divmod(int(minutes), 60)
        return f"{h:02d}h{m:02d}"


class AntifraudView(BaseAdminView):
    template_name = "administration/antifraud.html"
    active_nav = "antifraud"
    page_title = "Anti-fraude"
    breadcrumb = "Anti-fraude"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        try:
            from accounts.scoping import scope_queryset_by_company
            from antifraud.models import FraudAlert, FraudRule
            qs = FraudAlert.objects.select_related("rule", "site").order_by("-detected_at")
            # RBAC : restreint aux sites des filiales du user
            qs = scope_queryset_by_company(qs, self.request.user, "site__company")
            page_obj, page_qs = paginate(qs, self.request, per_page=25)
            ctx["page_obj"] = page_obj
            ctx["alerts"] = page_qs
            ctx["rules_active"] = FraudRule.objects.filter(is_active=True).count()
        except Exception:
            ctx["alerts"] = []
            ctx["rules_active"] = 0
        return ctx


class AuditView(BaseAdminView):
    template_name = "administration/audit.html"
    active_nav = "audit"
    page_title = "Audit & conformité"
    page_subtitle = "Journal d'audit, exports RGPD et politiques de rétention."
    breadcrumb = "Audit"
    permission_required = "audit.view"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        try:
            from audit.models import (AuditLog, ConformityRegister,
                                       DataExportRequest, LegalRetentionPolicy)
            tab = self.request.GET.get("tab", "logs")
            ctx["tab"] = tab if tab in ("logs", "exports", "policies", "registers") else "logs"

            q = (self.request.GET.get("q") or "").strip()
            ctx["q"] = q

            if ctx["tab"] == "logs":
                qs = AuditLog.objects.select_related("user").order_by("-timestamp")
                user_filter = self.request.GET.get("user", "")
                action_filter = self.request.GET.get("action", "")
                target_filter = self.request.GET.get("target", "")
                if q:
                    qs = qs.filter(
                        Q(action__icontains=q) | Q(target_model__icontains=q) |
                        Q(target_id__icontains=q) | Q(user__email__icontains=q)
                    )
                if user_filter:
                    qs = qs.filter(user__email__iexact=user_filter)
                if action_filter:
                    qs = qs.filter(action__iexact=action_filter)
                if target_filter:
                    qs = qs.filter(target_model__iexact=target_filter)
                page_obj, page_qs = paginate(qs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["logs"] = page_qs
                ctx["filter_user"] = user_filter
                ctx["filter_action"] = action_filter
                ctx["filter_target"] = target_filter
                # Choix pour les selects
                from django.db.models import Count as _Count
                ctx["distinct_actions"] = list(AuditLog.objects.values_list(
                    "action", flat=True).distinct().order_by("action")[:50])
                ctx["distinct_targets"] = list(AuditLog.objects.exclude(
                    target_model="").values_list(
                    "target_model", flat=True).distinct().order_by("target_model")[:50])
            elif ctx["tab"] == "exports":
                qs = (DataExportRequest.objects
                      .select_related("requested_by")
                      .order_by("-created_at"))
                page_obj, page_qs = paginate(qs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["exports"] = page_qs
            elif ctx["tab"] == "policies":
                qs = LegalRetentionPolicy.objects.order_by("-is_active", "target_model")
                page_obj, page_qs = paginate(qs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["policies"] = page_qs
            else:
                qs = (ConformityRegister.objects.select_related("site", "performed_by")
                      .order_by("-performed_at"))
                page_obj, page_qs = paginate(qs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["registers"] = page_qs

            ctx["logs_total"] = AuditLog.objects.count()
            ctx["exports_pending"] = DataExportRequest.objects.filter(status="pending").count()
            ctx["policies_active"] = LegalRetentionPolicy.objects.filter(is_active=True).count()
            ctx["registers_total"] = ConformityRegister.objects.count()
        except Exception:
            for k in ("logs", "exports", "policies", "registers"):
                ctx[k] = []
            for k in ("logs_total", "exports_pending", "policies_active",
                      "registers_total"):
                ctx[k] = 0
            ctx["page_obj"] = None
            ctx["tab"] = "logs"
        return ctx


class NotificationsView(BaseAdminView):
    template_name = "administration/notifications.html"
    active_nav = "notifications"
    page_title = "Notifications"
    page_subtitle = "Templates, notifications envoyées et préférences."
    breadcrumb = "Notifications"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        try:
            from notifications.models import (Notification, NotificationPreference,
                                                NotificationTemplate)
            tab = self.request.GET.get("tab", "sent")
            ctx["tab"] = tab if tab in ("sent", "templates", "preferences") else "sent"

            if ctx["tab"] == "sent":
                qs = (Notification.objects.select_related("recipient", "template")
                      .order_by("-created_at"))
                status = self.request.GET.get("status", "")
                if status in ("sent", "delivered", "failed", "read", "queued"):
                    qs = qs.filter(status=status)
                page_obj, page_qs = paginate(qs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["notifications"] = page_qs
                ctx["status"] = status
            elif ctx["tab"] == "templates":
                qs = NotificationTemplate.objects.order_by("code")
                page_obj, page_qs = paginate(qs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["templates"] = page_qs
            else:
                qs = (NotificationPreference.objects.select_related("user")
                      .order_by("user__email"))
                page_obj, page_qs = paginate(qs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["preferences"] = page_qs

            ctx["sent_total"] = Notification.objects.count()
            ctx["sent_failed"] = Notification.objects.filter(status="failed").count()
            ctx["sent_today"] = Notification.objects.filter(
                created_at__date__gte=__import__("django").utils.timezone.now().date(),
            ).count()
            ctx["templates_total"] = NotificationTemplate.objects.count()
        except Exception:
            for k in ("notifications", "templates", "preferences"):
                ctx[k] = []
            for k in ("sent_total", "sent_failed", "sent_today", "templates_total"):
                ctx[k] = 0
            ctx["page_obj"] = None
            ctx["tab"] = "sent"
        return ctx


class MobileSyncView(BaseAdminView):
    template_name = "administration/mobile.html"
    active_nav = "mobile"
    page_title = "Mobile & sync"
    page_subtitle = "Devices terrain, file offline et sessions de synchronisation."
    breadcrumb = "Mobile"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        try:
            from mobile_sync.models import (MobileBundle, MobileDevice,
                                              OfflineScanQueue, SyncSession)
            tab = self.request.GET.get("tab", "devices")
            ctx["tab"] = tab if tab in ("devices", "queue", "sessions", "bundles") else "devices"

            if ctx["tab"] == "devices":
                qs = (MobileDevice.objects.select_related("user", "site")
                      .order_by("-last_sync_at"))
                page_obj, page_qs = paginate(qs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["devices"] = page_qs
            elif ctx["tab"] == "queue":
                qs = (OfflineScanQueue.objects.select_related("device")
                      .order_by("-captured_at"))
                page_obj, page_qs = paginate(qs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["queue_items"] = page_qs
            elif ctx["tab"] == "sessions":
                qs = (SyncSession.objects.select_related("device")
                      .order_by("-started_at"))
                page_obj, page_qs = paginate(qs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["sessions"] = page_qs
            else:
                qs = MobileBundle.objects.order_by("-created_at")
                page_obj, page_qs = paginate(qs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["bundles"] = page_qs

            ctx["devices_total"] = MobileDevice.objects.count()
            ctx["devices_active"] = MobileDevice.objects.filter(status="active").count()
            ctx["queue_pending"] = OfflineScanQueue.objects.filter(status="queued").count()
            ctx["sessions_today"] = SyncSession.objects.filter(
                started_at__date=__import__("django").utils.timezone.now().date(),
            ).count()
        except Exception:
            for k in ("devices", "queue_items", "sessions", "bundles"):
                ctx[k] = []
            for k in ("devices_total", "devices_active", "queue_pending",
                      "sessions_today"):
                ctx[k] = 0
            ctx["page_obj"] = None
            ctx["tab"] = "devices"
        return ctx


class ReportsView(BaseAdminView):
    template_name = "administration/reports.html"
    active_nav = "reports"
    page_title = "Rapports & KPI"
    page_subtitle = "Rapports configurables, exécutions et planifications."
    breadcrumb = "Rapports"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        try:
            from reports.models import (KPISnapshot, Report, ReportRun,
                                          ReportSchedule)
            tab = self.request.GET.get("tab", "reports")
            ctx["tab"] = tab if tab in ("reports", "runs", "schedules", "kpis") else "reports"

            if ctx["tab"] == "reports":
                qs = Report.objects.order_by("-is_active", "name")
                page_obj, page_qs = paginate(qs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["reports"] = page_qs
            elif ctx["tab"] == "runs":
                qs = (ReportRun.objects.select_related("report", "requested_by")
                      .order_by("-created_at"))
                page_obj, page_qs = paginate(qs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["runs"] = page_qs
            elif ctx["tab"] == "schedules":
                qs = (ReportSchedule.objects.select_related("report")
                      .order_by("-is_active", "report__name"))
                page_obj, page_qs = paginate(qs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["schedules"] = page_qs
            else:
                qs = (KPISnapshot.objects.select_related("site")
                      .order_by("-date"))
                page_obj, page_qs = paginate(qs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["kpis"] = page_qs

            ctx["reports_total"] = Report.objects.count()
            ctx["reports_active"] = Report.objects.filter(is_active=True).count()
            ctx["runs_pending"] = ReportRun.objects.filter(status="pending").count()
            ctx["schedules_active"] = ReportSchedule.objects.filter(is_active=True).count()

            # ── Insights IA / analytics déterministe ───────────────
            period = self.request.GET.get("period", "week")
            start = self.request.GET.get("start", "")
            end = self.request.GET.get("end", "")
            ctx["period"] = period if period in ("day", "week", "month",
                                                  "quarter", "year", "custom") else "week"
            ctx["period_start"] = start
            ctx["period_end"] = end
            try:
                from reports.insights import (compute_attendance_breakdown,
                                                compute_insights,
                                                executive_summary,
                                                resolve_period)
                insights = compute_insights(period=ctx["period"], start=start, end=end)
                ctx["insights"] = insights
                ctx["executive_summary"] = executive_summary(insights) if insights else ""
                ctx["attendance"] = compute_attendance_breakdown(
                    period=ctx["period"], start=start, end=end)
                _, _, _, _, ctx["period_label"] = resolve_period(
                    ctx["period"], start, end)
            except Exception:
                ctx["insights"] = []
                ctx["executive_summary"] = ""
                ctx["attendance"] = {"by_company": [], "by_site": [], "label": ""}
                ctx["period_label"] = ""
        except Exception:
            for k in ("reports", "runs", "schedules", "kpis"):
                ctx[k] = []
            for k in ("reports_total", "reports_active", "runs_pending",
                      "schedules_active"):
                ctx[k] = 0
            ctx["page_obj"] = None
            ctx["insights"] = []
            ctx["executive_summary"] = ""
            ctx["tab"] = "reports"
        return ctx


class AIAssistantView(BaseAdminView):
    template_name = "administration/ai.html"
    active_nav = "ai"
    page_title = "Assistant IA"
    page_subtitle = "Conversations, prompts système et appels d'outils."
    breadcrumb = "IA"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        try:
            from ai_assistant.models import (AIConversation, AIMessage,
                                               AIPromptTemplate, AIToolCall)
            tab = self.request.GET.get("tab", "conversations")
            ctx["tab"] = tab if tab in ("conversations", "templates", "tools") else "conversations"

            if ctx["tab"] == "conversations":
                qs = (AIConversation.objects.select_related("user", "site")
                      .order_by("-last_activity_at"))
                page_obj, page_qs = paginate(qs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["conversations"] = page_qs
            elif ctx["tab"] == "templates":
                qs = AIPromptTemplate.objects.order_by("-is_active", "code")
                page_obj, page_qs = paginate(qs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["templates"] = page_qs
            else:
                qs = AIToolCall.objects.select_related("message").order_by("-created_at")
                page_obj, page_qs = paginate(qs, self.request, per_page=25)
                ctx["page_obj"] = page_obj
                ctx["tools"] = page_qs

            ctx["conversations_total"] = AIConversation.objects.count()
            ctx["messages_total"] = AIMessage.objects.count()
            ctx["templates_active"] = AIPromptTemplate.objects.filter(is_active=True).count()
            ctx["tool_calls_total"] = AIToolCall.objects.count()
        except Exception:
            for k in ("conversations", "templates", "tools"):
                ctx[k] = []
            for k in ("conversations_total", "messages_total",
                      "templates_active", "tool_calls_total"):
                ctx[k] = 0
            ctx["page_obj"] = None
            ctx["tab"] = "conversations"
        return ctx


class FaceRecognitionTestView(BaseAdminView):
    template_name = "administration/face_test.html"
    active_nav = "employees"
    page_title = "Test reconnaissance faciale"
    breadcrumb = "Employés · Test face"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        try:
            from employees.models import Employee, FaceProfile
            ctx["employees_for_enroll"] = list(
                Employee.objects.filter(status="active")
                .order_by("last_name", "first_name")
                .values("id", "matricule", "first_name", "last_name")[:500]
            )
            from django.db.models import Sum
            ctx["face_profiles_count"] = FaceProfile.objects.filter(is_active=True).count()
            ctx["face_matches_total"] = (
                FaceProfile.objects.filter(is_active=True)
                .aggregate(t=Sum("match_count"))["t"] or 0
            )
        except Exception:
            logger.warning("FaceRecognitionTestView : récupération employés/profils échouée", exc_info=True)
            ctx["employees_for_enroll"] = []
            ctx["face_profiles_count"] = 0
            ctx["face_matches_total"] = 0
        return ctx


class SettingsView(BaseAdminView):
    template_name = "administration/settings.html"
    active_nav = "settings"
    page_title = "Paramètres"
    breadcrumb = "Paramètres"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        try:
            from core.services import get_kaydan_tenant
            ctx["tenant"] = get_kaydan_tenant()
        except Exception:
            ctx["tenant"] = None
        return ctx


class AccountsView(BaseAdminView):
    template_name = "administration/accounts.html"
    active_nav = "accounts"
    page_title = "Utilisateurs & rôles"
    breadcrumb = "Utilisateurs"
    permission_required = "accounts.view"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        try:
            from django.contrib.auth import get_user_model

            from accounts.models import APIKey, Role
            User = get_user_model()
            qs = User.objects.order_by("-date_joined")
            page_obj, page_qs = paginate(qs, self.request, per_page=25)
            ctx["page_obj"] = page_obj
            ctx["users"] = page_qs
            ctx["users_total"] = User.objects.count()
            ctx["users_active"] = User.objects.filter(is_active=True).count()
            ctx["users_staff"] = User.objects.filter(is_staff=True).count()
            ctx["roles"] = Role.objects.all()[:20]
            ctx["roles_count"] = Role.objects.count()
            ctx["api_keys"] = APIKey.objects.order_by("-created_at")[:20]
            ctx["api_keys_count"] = APIKey.objects.count()
        except Exception:
            ctx["users"] = []
            ctx["users_total"] = 0
            ctx["users_active"] = 0
            ctx["users_staff"] = 0
            ctx["roles"] = []
            ctx["roles_count"] = 0
            ctx["api_keys"] = []
            ctx["api_keys_count"] = 0
        return ctx


class CompaniesView(BaseAdminView):
    """Liste des filiales KAYDAN."""
    template_name = "administration/companies.html"
    active_nav = "companies"
    page_title = "Filiales KAYDAN"
    breadcrumb = "Filiales"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        try:
            from core.models import Company
            qs = Company.objects.all().order_by("-is_active", "name")
            page_obj, page_qs = paginate(qs, self.request, per_page=25)
            ctx["page_obj"] = page_obj
            ctx["companies"] = page_qs
            ctx["total"] = Company.objects.count()
            ctx["active_count"] = Company.objects.filter(is_active=True).count()
        except Exception:
            ctx["companies"] = []
            ctx["total"] = 0
            ctx["active_count"] = 0
        return ctx


# =============================================================================
# Cartographie (Leaflet)
# =============================================================================
class MapView(BaseAdminView):
    template_name = "administration/map.html"
    active_nav = "map"
    page_title = "Cartographie du contrôle d'accès"
    page_subtitle = "Sites, zones, checkpoints, présence en temps réel et plan d'évacuation."
    breadcrumb = "Cartographie"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        try:
            from sites.models import Site
            ctx["sites"] = Site.objects.filter(status="active").order_by("name")
        except Exception:
            ctx["sites"] = []
        return ctx


class MapDataAPIView(View):
    """GET /map/data/ — JSON consolidé pour Leaflet."""

    def get(self, request):
        from datetime import timedelta

        from django.utils import timezone

        from sites.models import Checkpoint, Site, Zone

        site_id = request.GET.get("site", "")
        holder_kind = request.GET.get("holder_kind", "")
        zone_id = request.GET.get("zone", "")
        event_filter = request.GET.get("event", "")
        emergency = request.GET.get("emergency") == "1"

        try:
            sites_qs = Site.objects.filter(status="active")
            if site_id and site_id.isdigit():
                sites_qs = sites_qs.filter(id=int(site_id))
            sites = list(sites_qs)

            site_coords = {
                s.id: (float(s.latitude), float(s.longitude))
                for s in sites
                if s.latitude is not None and s.longitude is not None
            }
            default_center = (5.345317, -4.024429)

            sites_payload = []
            for s in sites:
                lat = float(s.latitude) if s.latitude is not None else default_center[0]
                lng = float(s.longitude) if s.longitude is not None else default_center[1]
                geofence = s.geofence if isinstance(s.geofence, dict) else {}
                polygon = None
                if geofence and geofence.get("type") == "Polygon":
                    polygon = geofence.get("coordinates")
                sites_payload.append({
                    "id": s.id, "code": s.code, "name": s.name,
                    "type": s.type, "type_label": s.get_type_display(),
                    "lat": lat, "lng": lng,
                    "polygon": polygon,
                    "risk_level": s.risk_level,
                })

            zones_qs = Zone.objects.select_related("site")
            if site_id and site_id.isdigit():
                zones_qs = zones_qs.filter(site_id=int(site_id))
            if zone_id and zone_id.isdigit():
                zones_qs = zones_qs.filter(id=int(zone_id))
            zones = list(zones_qs)

            zones_payload = []
            for idx, z in enumerate(zones):
                lat0, lng0 = site_coords.get(z.site_id, default_center)
                row = idx // 4
                col = idx % 4
                dy = (row - 1) * 0.0014
                dx = (col - 1.5) * 0.0014
                cy, cx = lat0 + dy, lng0 + dx
                size = 0.00045
                ring = [[cx - size, cy - size], [cx + size, cy - size],
                        [cx + size, cy + size], [cx - size, cy + size],
                        [cx - size, cy - size]]
                zone_type = "office"
                lname = z.name.lower()
                if "stock" in lname or "entrepot" in lname or "magasin" in lname:
                    zone_type = "storage"
                elif "chant" in lname:
                    zone_type = "construction"
                elif "rassem" in lname or "evac" in lname or "muster" in lname:
                    zone_type = "muster"
                elif "interd" in lname or z.is_restricted:
                    zone_type = "restricted"
                elif "entree" in lname or "sortie" in lname or "porti" in lname:
                    zone_type = "gate"
                elif "sensible" in lname or "danger" in lname:
                    zone_type = "sensitive"
                zones_payload.append({
                    "id": z.id, "site_id": z.site_id,
                    "name": z.name, "code": z.code,
                    "type": zone_type,
                    "is_restricted": bool(z.is_restricted),
                    "polygon": [ring],
                    "center": [cy, cx],
                })

            cp_qs = Checkpoint.objects.select_related("site", "zone").filter(is_active=True)
            if site_id and site_id.isdigit():
                cp_qs = cp_qs.filter(site_id=int(site_id))
            cp_payload = []
            import math
            cp_total = max(1, cp_qs.count())
            for idx, c in enumerate(cp_qs):
                lat0, lng0 = site_coords.get(c.site_id, default_center)
                angle = (idx * 360 / cp_total) * math.pi / 180
                radius = 0.0011
                lat = lat0 + radius * math.cos(angle)
                lng = lng0 + radius * math.sin(angle)
                cp_payload.append({
                    "id": c.id, "site_id": c.site_id, "zone_id": c.zone_id,
                    "name": c.name, "type": c.type, "type_label": c.get_type_display(),
                    "method": c.method, "method_label": c.get_method_display(),
                    "lat": lat, "lng": lng,
                })

            from access_control.models import AccessEvent
            now = timezone.now()
            since = now - timedelta(hours=12 if not emergency else 24)
            ev_qs = AccessEvent.objects.filter(timestamp__gte=since).select_related("site")
            if site_id and site_id.isdigit():
                ev_qs = ev_qs.filter(site_id=int(site_id))
            if holder_kind in ("employee", "worker", "visitor"):
                ev_qs = ev_qs.filter(holder_kind=holder_kind)
            if event_filter in ("granted", "denied", "review"):
                ev_qs = ev_qs.filter(decision=event_filter)

            latest_by_holder = {}
            for ev in ev_qs.order_by("-timestamp").iterator():
                key = (ev.holder_kind, ev.holder_object_id)
                if not key[0] or not key[1]:
                    continue
                if key not in latest_by_holder:
                    latest_by_holder[key] = ev

            people_events = list(latest_by_holder.values())
            people_events = _annotate_events_with_holders(people_events)

            people_payload = []
            anomalies = []
            for ev in people_events:
                if not ev.holder_obj:
                    continue
                lat0, lng0 = site_coords.get(ev.site_id, default_center)
                import hashlib
                seed = int(hashlib.md5(
                    f"{ev.holder_kind}-{ev.holder_object_id}".encode()
                ).hexdigest()[:8], 16)
                dx = ((seed % 200) - 100) / 100000.0
                dy = ((seed >> 8) % 200 - 100) / 100000.0

                anom = []
                if ev.decision == "denied":
                    anom.append({"kind": "access_denied", "label": ev.denial_reason or "Refus accès"})
                if ev.holder_kind == "worker" and not ev.badge_uid:
                    anom.append({"kind": "no_badge", "label": "Aucun badge scanné"})

                photo_url = ""
                holder = ev.holder_obj
                if getattr(holder, "photo", None) and holder.photo.name:
                    try:
                        photo_url = holder.photo.url
                    except ValueError:
                        photo_url = ""

                people_payload.append({
                    "kind": ev.holder_kind,
                    "id": ev.holder_object_id,
                    "name": ev.holder_label,
                    "matricule": getattr(holder, "matricule", "") or "",
                    "photo": photo_url,
                    "lat": lat0 + dx, "lng": lng0 + dy,
                    "last_site": ev.site.name if ev.site else "",
                    "last_site_id": ev.site_id,
                    "last_seen": ev.timestamp.isoformat(),
                    "decision": ev.decision,
                    "badge_uid": ev.badge_uid,
                    "anomalies": anom,
                })
                if anom:
                    for a in anom:
                        anomalies.append({
                            "person": ev.holder_label,
                            "kind": ev.holder_kind,
                            "site": ev.site.name if ev.site else "",
                            "anomaly": a["kind"],
                            "label": a["label"],
                            "timestamp": ev.timestamp.isoformat(),
                            "lat": lat0 + dx, "lng": lng0 + dy,
                        })

            try:
                from antifraud.models import FraudAlert
                fa_qs = FraudAlert.objects.filter(
                    raised_at__gte=since, status__in=("open", "escalated"),
                ).select_related("rule", "site")[:50]
                for a in fa_qs:
                    lat, lng = site_coords.get(a.site_id, default_center)
                    anomalies.append({
                        "person": "—",
                        "kind": "fraud_alert",
                        "site": a.site.name if a.site else "",
                        "anomaly": a.rule.code if a.rule else "alert",
                        "label": a.message or (a.rule.name if a.rule else "Alerte fraude"),
                        "timestamp": a.raised_at.isoformat(),
                        "lat": lat, "lng": lng,
                    })
            except Exception:
                logger.debug("Cartographie: anomalies fraude — point ignoré", exc_info=True)

            from collections import Counter
            kind_counts = Counter(p["kind"] for p in people_payload)
            site_counts = Counter(p["last_site_id"] for p in people_payload)

            return JsonResponse({
                "sites": sites_payload,
                "zones": zones_payload,
                "checkpoints": cp_payload,
                "people": people_payload,
                "anomalies": anomalies,
                "stats": {
                    "total_present": len(people_payload),
                    "employees": kind_counts.get("employee", 0),
                    "workers": kind_counts.get("worker", 0),
                    "visitors": kind_counts.get("visitor", 0),
                    "anomalies_count": len(anomalies),
                    "by_site": [
                        {"site_id": sid, "site_name": next(
                            (s["name"] for s in sites_payload if s["id"] == sid), "—"),
                         "count": cnt}
                        for sid, cnt in site_counts.most_common()
                    ],
                },
                "default_center": list(default_center),
                "now": now.isoformat(),
            })

        except Exception as e:
            import logging
            logging.getLogger(__name__).exception("MapDataAPIView failed")
            return JsonResponse({"error": str(e)}, status=500)


# =============================================================================
# Export Excel — Flux temps réel
# =============================================================================
class RealtimeExportView(View):
    """GET /realtime/export/?<filtres> — export xlsx."""
    MAX_ROWS = 50_000

    def get(self, request):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            return HttpResponse(
                "openpyxl non installé — pip install openpyxl",
                status=500, content_type="text/plain",
            )

        from datetime import datetime as _dt

        from django.utils import timezone

        qs = _realtime_filtered_qs(request)[:self.MAX_ROWS]
        events = _annotate_events_with_holders(list(qs))

        wb = Workbook()
        ws = wb.active
        ws.title = "Flux temps réel"

        headers = [
            "Horodatage", "Site", "Type porteur", "Nom", "Matricule",
            "Filiale / Sous-traitant", "Badge UID", "Méthode",
            "Décision", "Motif refus", "Score visage",
        ]
        ws.append(headers)
        header_fill = PatternFill(start_color="0B1B33", end_color="0B1B33", fill_type="solid")
        header_font = Font(bold=True, color="F26B1F", size=11)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        kind_label = {"employee": "Employé", "worker": "Ouvrier",
                       "visitor": "Visiteur", "unknown": "Inconnu"}
        decision_label = {"granted": "Autorisé", "denied": "Refusé", "review": "À vérifier"}

        for e in events:
            holder = getattr(e, "holder_obj", None)
            name = ""
            matricule = ""
            company = ""
            if holder:
                name = f"{getattr(holder, 'first_name', '')} {getattr(holder, 'last_name', '')}".strip()
                matricule = getattr(holder, "matricule", "") or ""
                if e.holder_kind == "employee":
                    company = getattr(getattr(holder, "company", None), "name", "") or ""
                elif e.holder_kind == "worker":
                    company = getattr(getattr(holder, "subcontractor", None), "name", "") or "Direct"
                elif e.holder_kind == "visitor":
                    company = getattr(holder, "company", "") or ""

            ws.append([
                timezone.localtime(e.timestamp).strftime("%d/%m/%Y %H:%M:%S"),
                e.site.name if e.site else "",
                kind_label.get(e.holder_kind, e.holder_kind or ""),
                name,
                matricule,
                company,
                e.badge_uid or "",
                (e.method or "").upper(),
                decision_label.get(e.decision, e.decision),
                e.denial_reason or "",
                f"{e.face_match_score:.2f}" if getattr(e, "face_match_score", None) else "",
            ])

        widths = [22, 26, 14, 28, 16, 22, 22, 12, 14, 26, 12]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        applied = []
        for k in ("period", "holder_kind", "site", "decision", "individual", "start", "end"):
            if request.GET.get(k):
                applied.append(f"{k}={request.GET.get(k)}")
        ws.append([])
        ws.append([
            "Export généré le",
            timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M"),
            "Filtres :", " · ".join(applied) or "aucun",
            "Total lignes", len(events),
            f"(plafond {self.MAX_ROWS})" if len(events) >= self.MAX_ROWS else "",
            "", "", "", "",
        ])

        import io
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"flux_temps_reel_{_dt.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response = HttpResponse(
            buf.getvalue(),
            content_type=("application/vnd.openxmlformats-officedocument."
                          "spreadsheetml.sheet"),
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


# =============================================================================
# Gestion des utilisateurs (vues natives KAYDAN — pas de Django admin)
# =============================================================================
class _UserCRUDMixin:
    active_nav = "accounts"
    list_url_name = "admin-accounts"
    entity_label = "Utilisateur"
    entity_label_plural = "Utilisateurs"
    url_key = "user"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update({
            "active_nav": self.active_nav,
            "list_url_name": self.list_url_name,
            "entity_label": self.entity_label,
            "entity_label_plural": self.entity_label_plural,
            "url_key": self.url_key,
            "open_alerts_count": 0,
        })
        return ctx


class UserCreateView(_UserCRUDMixin, CreateView):
    template_name = "administration/_form.html"
    success_url = reverse_lazy("admin-accounts")
    page_title = "Nouvel utilisateur"
    breadcrumb = "Utilisateurs · Nouveau"

    def get_form_class(self):
        from .forms import UserCreateForm
        return UserCreateForm

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["page_title"] = self.page_title
        ctx["breadcrumb"] = self.breadcrumb
        return ctx

    def form_valid(self, form):
        response = super().form_valid(form)
        dj_messages.success(self.request,
            f"Utilisateur {self.object.email} créé avec succès.")
        return response


class UserUpdateView(_UserCRUDMixin, UpdateView):
    template_name = "administration/_form.html"
    breadcrumb = "Utilisateurs · Modifier"

    def get_queryset(self):
        from django.contrib.auth import get_user_model
        return get_user_model().objects.all()

    def get_form_class(self):
        from .forms import UserUpdateForm
        return UserUpdateForm

    def get_success_url(self):
        return reverse("admin-user-detail", args=[self.object.pk])

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["page_title"] = f"Modifier · {self.object.get_full_name() or self.object.email}"
        ctx["breadcrumb"] = self.breadcrumb
        return ctx

    def form_valid(self, form):
        response = super().form_valid(form)
        dj_messages.success(self.request, "Utilisateur mis à jour.")
        return response


class UserDetailView(_UserCRUDMixin, DetailView):
    template_name = "administration/user_detail.html"

    def get_queryset(self):
        from django.contrib.auth import get_user_model
        return get_user_model().objects.all()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        u = self.object
        ctx["page_title"] = f"Utilisateur · {u.get_full_name() or u.email}"
        ctx["breadcrumb"] = f"Utilisateurs · {u.email}"
        try:
            from accounts.models import RoleAssignment
            ctx["assignments"] = list(
                RoleAssignment.objects.filter(user=u)
                .select_related("role", "site").order_by("role__name")
            )
            perms = set()
            for a in ctx["assignments"]:
                perms.update(a.role.permissions.values_list("code", flat=True))
            ctx["aggregated_permissions"] = sorted(perms)
        except Exception:
            ctx["assignments"] = []
            ctx["aggregated_permissions"] = []
        try:
            ctx["employee_profile"] = u.employee_profile
        except Exception:
            ctx["employee_profile"] = None
        try:
            from accounts.models import LoginAttempt, UserSession
            ctx["sessions"] = list(
                UserSession.objects.filter(user=u).order_by("-last_activity_at")[:10]
            )
            ctx["login_attempts"] = list(
                LoginAttempt.objects.filter(email__iexact=u.email).order_by("-created_at")[:10]
            )
        except Exception:
            ctx["sessions"] = []
            ctx["login_attempts"] = []

        ctx["update_url"] = reverse("admin-user-update", args=[u.pk])
        ctx["password_url"] = reverse("admin-user-password", args=[u.pk])
        ctx["toggle_url"] = reverse("admin-user-toggle", args=[u.pk])
        return ctx


class UserPasswordView(_UserCRUDMixin, FormView):
    """Réinitialise le mot de passe d'un utilisateur."""
    template_name = "administration/user_password.html"

    def get_form_class(self):
        from .forms import UserPasswordForm
        return UserPasswordForm

    def get_user(self):
        from django.contrib.auth import get_user_model
        return get_object_or_404(get_user_model(), pk=self.kwargs["pk"])

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["target_user"] = self.get_user()
        ctx["page_title"] = f"Réinitialiser le mot de passe · {ctx['target_user'].email}"
        ctx["breadcrumb"] = "Utilisateurs · Mot de passe"
        return ctx

    def form_valid(self, form):
        u = self.get_user()
        u.set_password(form.cleaned_data["password"])
        u.save(update_fields=["password"])
        dj_messages.success(self.request, f"Mot de passe mis à jour pour {u.email}.")
        return redirect("admin-user-detail", pk=u.pk)


class UserToggleActiveView(_UserCRUDMixin, View):
    """POST → bascule is_active."""

    def post(self, request, pk):
        from django.contrib.auth import get_user_model
        u = get_object_or_404(get_user_model(), pk=pk)
        u.is_active = not u.is_active
        u.save(update_fields=["is_active"])
        verb = "réactivé" if u.is_active else "désactivé"
        dj_messages.success(request, f"Compte {u.email} {verb}.")
        return redirect("admin-user-detail", pk=pk)


class UserDeleteView(_UserCRUDMixin, View):
    """GET → confirme · POST → désactive (soft delete) le user.

    On ne supprime JAMAIS un User physiquement (RGPD + intégrité référentielle
    sur AccessEvent, AuditLog…). On désactive et anonymise.
    """
    template_name = "administration/_confirm_delete.html"

    def get(self, request, pk):
        from django.contrib.auth import get_user_model
        from django.shortcuts import render
        u = get_object_or_404(get_user_model(), pk=pk)
        return render(request, self.template_name, {
            "object": u, "page_title": f"Désactiver · {u.email}",
            "entity_label": "Utilisateur", "entity_label_plural": "Utilisateurs",
            "list_url_name": "admin-accounts", "url_key": "user",
            "active_nav": "accounts",
            "delete_warning": ("Pour des raisons d'audit RGPD, le compte sera "
                               "désactivé (is_active=False) plutôt que supprimé. "
                               "L'historique d'accès et d'audit est préservé."),
        })

    def post(self, request, pk):
        from django.contrib.auth import get_user_model
        u = get_object_or_404(get_user_model(), pk=pk)
        if u == request.user:
            dj_messages.error(request, "Vous ne pouvez pas désactiver votre propre compte.")
            return redirect("admin-user-detail", pk=pk)
        u.is_active = False
        u.save(update_fields=["is_active"])
        dj_messages.success(request, f"Compte {u.email} désactivé.")
        return redirect("admin-accounts")


# Rôles
class RoleListView(_UserCRUDMixin, BaseAdminView):
    template_name = "administration/roles.html"
    page_title = "Rôles & permissions"
    breadcrumb = "Rôles"
    entity_label = "Rôle"
    entity_label_plural = "Rôles"
    url_key = "role"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        try:
            from accounts.models import Role
            ctx["roles"] = list(
                Role.objects.prefetch_related("permissions").order_by("name")
            )
            for r in ctx["roles"]:
                r.user_count = r.assignments.values("user").distinct().count()
        except Exception:
            ctx["roles"] = []
        return ctx


class RoleCreateView(_UserCRUDMixin, CreateView):
    template_name = "administration/_form.html"
    page_title = "Nouveau rôle"
    entity_label = "Rôle"
    url_key = "role"

    def get_form_class(self):
        from .forms import RoleForm
        return RoleForm

    def get_success_url(self):
        return reverse("admin-roles")


class RoleUpdateView(_UserCRUDMixin, UpdateView):
    template_name = "administration/_form.html"
    entity_label = "Rôle"
    url_key = "role"

    def get_queryset(self):
        from accounts.models import Role
        return Role.objects.all()

    def get_form_class(self):
        from .forms import RoleForm
        return RoleForm

    def get_success_url(self):
        return reverse("admin-roles")


# ===========================================================================
# Gestion des APIKeys IoT — vue dédiée car le secret brut n'est visible qu'une
# seule fois après création (puis seul `secret_hash` reste en base).
# ===========================================================================
class APIKeyListView(BaseAdminView):
    """Liste des clés API IoT — utilisées pour signer les requêtes HMAC."""
    template_name = "administration/api_keys.html"
    active_nav = "accounts"
    page_title = "Clés API IoT"
    breadcrumb = "Utilisateurs · Clés API"
    permission_required = "apikeys.manage"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        try:
            from accounts.models import APIKey
            qs = (APIKey.objects.select_related("tenant", "site")
                  .order_by("-is_active", "-created_at"))
            from .views import paginate
            page_obj, page_qs = paginate(qs, self.request, per_page=25)
            ctx["page_obj"] = page_obj
            ctx["api_keys"] = page_qs
            ctx["total"] = APIKey.objects.count()
            ctx["active"] = APIKey.objects.filter(is_active=True, revoked_at__isnull=True).count()
            ctx["revoked"] = APIKey.objects.filter(revoked_at__isnull=False).count()
        except Exception:
            ctx["api_keys"] = []
            ctx["page_obj"] = None
            ctx["total"] = 0
            ctx["active"] = 0
            ctx["revoked"] = 0
        # Récupère le secret one-shot stocké en session (si vient d'une création)
        ctx["new_key_secret"] = self.request.session.pop("new_api_key_secret", None)
        ctx["new_key_id"] = self.request.session.pop("new_api_key_public_id", None)
        return ctx


class APIKeyCreateView(BaseAdminView):
    """Création d'une APIKey — affiche le secret brut UNE seule fois."""
    template_name = "administration/_form.html"
    active_nav = "accounts"
    page_title = "Nouvelle clé API IoT"
    breadcrumb = "Utilisateurs · Clés API · Nouvelle"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from .forms import APIKeyForm
        ctx["form"] = kwargs.get("form") or APIKeyForm()
        ctx["entity_label"] = "Clé API"
        ctx["entity_label_plural"] = "Clés API"
        ctx["list_url_name"] = "admin-api-keys"
        return ctx

    def post(self, request, *args, **kwargs):
        import hashlib
        import secrets

        from .forms import APIKeyForm
        form = APIKeyForm(request.POST)
        if not form.is_valid():
            return self.render_to_response(self.get_context_data(form=form))

        from accounts.models import APIKey
        from core.services import get_kaydan_tenant
        instance = form.save(commit=False)
        if not instance.tenant_id:
            try:
                instance.tenant = get_kaydan_tenant()
            except Exception:
                logger.warning("Auto-tenant KAYDAN sur APIKey échoué", exc_info=True)
        # Génération secret + public_id ; on stocke uniquement le hash
        public_id = secrets.token_urlsafe(12)[:32]
        raw_secret = APIKey.generate_secret()
        instance.public_id = public_id
        instance.secret_hash = hashlib.sha256(raw_secret.encode()).hexdigest()
        instance.save()

        # Le secret brut transite UNE seule fois via la session puis disparaît
        request.session["new_api_key_secret"] = raw_secret
        request.session["new_api_key_public_id"] = public_id
        dj_messages.success(request,
            f"Clé API « {instance.name} » créée. Notez le secret immédiatement, "
            "il ne sera plus affiché.")
        return redirect("admin-api-keys")


class APIKeyRevokeView(View):
    """POST → révoque une APIKey (is_active=False + revoked_at=now)."""

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect("admin-login")
        from accounts.rbac import user_has_permission
        if not user_has_permission(request.user, "apikeys.manage"):
            dj_messages.error(request, "Permission refusée.")
            return redirect("admin-api-keys")
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, pk):
        from django.utils import timezone

        from accounts.models import APIKey
        key = get_object_or_404(APIKey, pk=pk)
        key.is_active = False
        key.revoked_at = timezone.now()
        key.save(update_fields=["is_active", "revoked_at"])
        dj_messages.success(request, f"Clé API « {key.name} » révoquée.")
        return redirect("admin-api-keys")


# ===========================================================================
# Mini-API notifications (consommée par le dropdown de la topbar)
# ===========================================================================
class NotificationUnreadCountView(View):
    """GET /api/notifications/unread/ → {count: N} pour l'utilisateur courant."""

    def get(self, request):
        if not request.user.is_authenticated:
            return JsonResponse({"count": 0})
        try:
            from notifications.models import Notification
            n = Notification.objects.filter(
                recipient=request.user,
                read_at__isnull=True,
            ).count()
            return JsonResponse({"count": n})
        except Exception:
            return JsonResponse({"count": 0})


class NotificationRecentView(View):
    """GET /api/notifications/recent/ → 20 dernières notifications du user."""

    def get(self, request):
        if not request.user.is_authenticated:
            return JsonResponse({"results": []})
        try:
            from django.utils import timezone

            from notifications.models import Notification
            notifs = list(Notification.objects.filter(
                recipient=request.user,
            ).order_by("-created_at")[:20])
            now = timezone.now()
            results = []
            for n in notifs:
                # severity inférée du payload (alert_id) ou de la rule
                severity = "info"
                payload = n.payload or {}
                if "alert_id" in payload:
                    try:
                        from antifraud.models import FraudAlert
                        a = FraudAlert.objects.filter(pk=payload["alert_id"]).first()
                        if a:
                            severity = a.severity or "medium"
                    except Exception:
                        logger.debug("Severity lookup FraudAlert %s échoué",
                                      payload.get("alert_id"), exc_info=True)

                delta = now - n.created_at
                if delta.total_seconds() < 60:
                    timeago = "à l'instant"
                elif delta.total_seconds() < 3600:
                    timeago = f"il y a {int(delta.total_seconds() // 60)} min"
                elif delta.total_seconds() < 86400:
                    timeago = f"il y a {int(delta.total_seconds() // 3600)} h"
                else:
                    timeago = f"il y a {delta.days} j"

                results.append({
                    "id": n.id,
                    "subject": n.subject or "Notification",
                    "body": (n.body or "")[:120],
                    "read": n.read_at is not None,
                    "severity": severity,
                    "timeago": timeago,
                })
            return JsonResponse({"results": results})
        except Exception:
            return JsonResponse({"results": []})


class NotificationMarkAllReadView(View):
    """POST /api/notifications/mark-all-read/ → marque toutes les notifs lues."""

    def post(self, request):
        if not request.user.is_authenticated:
            return JsonResponse({"ok": False}, status=401)
        try:
            from django.utils import timezone

            from notifications.models import Notification
            updated = Notification.objects.filter(
                recipient=request.user, read_at__isnull=True,
            ).update(read_at=timezone.now())
            return JsonResponse({"ok": True, "updated": updated})
        except Exception:
            return JsonResponse({"ok": False}, status=500)


# ===========================================================================
# Workflow visiteurs — actions métier sur VisitRequest
# ===========================================================================
class VisitRequestActionView(View):
    """POST /visit-requests/<pk>/action/<verb>/ → transitions de statut.

    Verbes : approve · reject · check_in · check_out · cancel
    Crée automatiquement les VisitLog et VisitorPass associés.
    """

    VALID_VERBS = ("approve", "reject", "check_in", "check_out", "cancel")

    def post(self, request, pk, verb):
        from datetime import timedelta

        from django.utils import timezone

        from visitors.models import VisitLog, VisitorPass, VisitRequest

        if verb not in self.VALID_VERBS:
            return redirect("admin-visitors")

        vr = get_object_or_404(VisitRequest, pk=pk)
        now = timezone.now()

        if verb == "approve" and vr.status in ("pending", "draft"):
            vr.status = "approved"
            vr.save(update_fields=["status"])
            try:
                from notifications.services import notify_visit_status_change
                notify_visit_status_change(vr, "approved")
            except Exception:
                logger.warning("Notification visite approuvée non envoyée (vr=%s)", vr.pk, exc_info=True)
            dj_messages.success(request, f"Demande de {vr.visitor} approuvée.")

        elif verb == "reject" and vr.status in ("pending", "draft"):
            vr.status = "rejected"
            vr.save(update_fields=["status"])
            dj_messages.warning(request, f"Demande de {vr.visitor} rejetée.")

        elif verb == "check_in" and vr.status in ("approved", "pending"):
            vr.status = "checked_in"
            vr.save(update_fields=["status"])
            VisitLog.objects.create(
                visit_request=vr, checked_in_at=now,
                checkin_user=request.user if request.user.is_authenticated else None,
            )
            # Émet automatiquement un VisitorPass valable pour la durée prévue
            VisitorPass.objects.get_or_create(
                visit_request=vr,
                defaults={
                    "type": "walk_in_pvc" if vr.mode == "walk_in" else "self_service_qr",
                    "valid_from": now,
                    "valid_until": now + timedelta(
                        minutes=vr.expected_duration_minutes or 60),
                },
            )
            try:
                from notifications.services import notify_visit_status_change
                notify_visit_status_change(vr, "checked_in")
            except Exception:
                logger.warning("Notification check-in visiteur non envoyée (vr=%s)", vr.pk, exc_info=True)
            dj_messages.success(request, f"{vr.visitor} a fait son check-in et reçu un pass.")

        elif verb == "check_out" and vr.status == "checked_in":
            log = vr.logs.filter(checked_out_at__isnull=True).first()
            if log:
                log.checked_out_at = now
                log.save(update_fields=["checked_out_at"])
            vr.status = "completed"
            vr.save(update_fields=["status"])
            # Révoque le pass associé (related_name="pass_card" sur VisitorPass)
            try:
                vp = getattr(vr, "pass_card", None)
                if vp and not vp.revoked_at:
                    vp.revoked_at = now
                    vp.save(update_fields=["revoked_at"])
            except Exception:
                import logging
                logging.getLogger(__name__).exception(
                    "check_out: échec révocation pass pour vr=%s", vr.id)
            dj_messages.success(request, f"{vr.visitor} a quitté le site.")

        elif verb == "cancel":
            vr.status = "cancelled"
            vr.save(update_fields=["status"])
            dj_messages.info(request, f"Demande {vr} annulée.")

        else:
            dj_messages.error(request,
                f"Action « {verb} » impossible sur le statut « {vr.status} ».")

        return redirect("admin-visitrequest-detail", pk=vr.pk)


class VisitCalendarICSView(View):
    """GET /visit-requests/<pk>/calendar.ics → fichier iCalendar."""

    def get(self, request, pk):
        from visitors.calendar_ics import visit_to_ics
        from visitors.models import VisitRequest
        vr = get_object_or_404(VisitRequest, pk=pk)
        ics = visit_to_ics(vr)
        response = HttpResponse(ics, content_type="text/calendar; charset=utf-8")
        response["Content-Disposition"] = (
            f'attachment; filename="visite-VR{vr.id}.ics"'
        )
        return response


class VisitorAddToWatchlistView(View):
    """POST /visitors/<pk>/watchlist/ → ajoute un visiteur à la liste rouge."""

    def post(self, request, pk):
        from visitors.models import Visitor, Watchlist
        v = get_object_or_404(Visitor, pk=pk)
        reason = request.POST.get("reason", "Ajout manuel")
        Watchlist.objects.create(
            tenant=v.tenant, visitor=v,
            full_name=f"{v.first_name} {v.last_name}",
            id_number=v.id_number, reason=reason, is_active=True,
        )
        dj_messages.warning(request, f"{v} ajouté à la liste rouge.")
        return redirect("admin-visitor-detail", pk=v.pk)


# ===========================================================================
# FraudAlert — actions métier (acquit, faux-positif, escalader, résoudre)
# ===========================================================================
class KshieldLoginView(View):
    """GET /auth/login/ → page login custom. POST → authentifie + redirige."""
    template_name = "administration/login.html"

    def get(self, request):
        from django.conf import settings
        from django.shortcuts import render
        if request.user.is_authenticated:
            from django.shortcuts import redirect
            return redirect(request.GET.get("next") or "admin-dashboard")
        return render(request, self.template_name, {
            "next": request.GET.get("next", ""),
            "sso_enabled": getattr(settings, "SSO_ENABLED", False),
        })

    def post(self, request):
        from django.contrib.auth import authenticate, login
        from django.shortcuts import redirect, render

        from accounts.models import LoginAttempt
        email = (request.POST.get("email") or "").strip()
        password = request.POST.get("password") or ""
        next_url = request.POST.get("next") or "admin-dashboard"

        user = authenticate(request, username=email, password=password)
        success = user is not None and user.is_active
        try:
            LoginAttempt.objects.create(
                email=email,
                ip=request.META.get("REMOTE_ADDR"),
                user_agent=(request.META.get("HTTP_USER_AGENT", "") or "")[:500],
                success=success,
                reason="" if success else "Identifiants invalides ou compte désactivé",
            )
        except Exception:
            # On ne bloque PAS le login si l'audit échoue, mais on trace pour le SIEM.
            logger.exception("LoginAttempt non persisté (email=%s, ok=%s)", email, success)

        if success:
            login(request, user)
            dj_messages.success(request, f"Bienvenue {user.get_full_name() or user.email}.")
            from accounts.rbac import invalidate_user_perms
            invalidate_user_perms(user.pk)
            try:
                return redirect(next_url)
            except Exception:
                logger.debug("Redirect next_url=%r invalide, fallback dashboard", next_url, exc_info=True)
                return redirect("admin-dashboard")

        from django.conf import settings
        return render(request, self.template_name, {
            "error": "Email ou mot de passe invalide.",
            "email": email, "next": next_url,
            "sso_enabled": getattr(settings, "SSO_ENABLED", False),
        })


class KshieldLogoutView(View):
    """GET ou POST /auth/logout/ → déconnecte + redirige login."""
    def post(self, request):
        from django.contrib.auth import logout
        from django.shortcuts import redirect
        logout(request)
        dj_messages.info(request, "Vous avez été déconnecté.")
        return redirect("admin-login")
    get = post


class CSVImportView(BaseAdminView):
    """Import CSV en masse pour Employee ou Worker.

    GET → page d'upload + format attendu.
    POST → parse le CSV, valide, crée les enregistrements en bulk.
    """
    template_name = "administration/csv_import.html"

    KIND_MAP = {
        "employees": {
            "label": "Employés",
            "active_nav": "employees",
            "fields": ["matricule", "first_name", "last_name", "email",
                       "phone", "contract_type", "status", "work_location"],
            "required": ["matricule", "first_name", "last_name"],
            "back_url": "admin-employees",
        },
        "workers": {
            "label": "Ouvriers",
            "active_nav": "workers",
            "fields": ["matricule", "first_name", "last_name", "phone",
                       "trade_code", "subcontractor_code", "status"],
            "required": ["matricule", "first_name", "last_name"],
            "back_url": "admin-workers",
        },
    }

    def get_kind(self):
        kind = self.kwargs.get("kind", "employees")
        return kind if kind in self.KIND_MAP else "employees"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        kind = self.get_kind()
        meta = self.KIND_MAP[kind]
        ctx["kind"] = kind
        ctx["meta"] = meta
        ctx["page_title"] = f"Import CSV — {meta['label']}"
        ctx["active_nav"] = meta["active_nav"]
        ctx["breadcrumb"] = f"{meta['label']} · Import CSV"
        ctx["template_csv"] = ",".join(meta["fields"])
        return ctx

    def post(self, request, kind):
        import csv
        import io
        meta = self.KIND_MAP.get(kind)
        if not meta:
            dj_messages.error(request, "Type d'import invalide.")
            return redirect("admin-dashboard")

        f = request.FILES.get("file")
        if not f:
            dj_messages.error(request, "Aucun fichier reçu.")
            return redirect("admin-csv-import", kind=kind)

        try:
            text = f.read().decode("utf-8-sig")
        except UnicodeDecodeError:
            dj_messages.error(request,
                "Fichier non UTF-8. Convertissez et réessayez.")
            return redirect("admin-csv-import", kind=kind)

        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            dj_messages.error(request, "Fichier vide ou sans en-têtes.")
            return redirect("admin-csv-import", kind=kind)

        # Vérifie les colonnes requises
        missing = set(meta["required"]) - set(reader.fieldnames)
        if missing:
            dj_messages.error(request,
                f"Colonnes manquantes : {', '.join(missing)}")
            return redirect("admin-csv-import", kind=kind)

        from core.services import get_kaydan_tenant
        tenant = get_kaydan_tenant()

        created, skipped, errors = 0, 0, []
        rows = list(reader)
        for i, row in enumerate(rows, start=2):  # ligne 2 = première data
            try:
                if kind == "employees":
                    n = self._create_employee(row, tenant)
                else:
                    n = self._create_worker(row, tenant)
                if n:
                    created += 1
                else:
                    skipped += 1
            except Exception as exc:
                errors.append(f"Ligne {i} : {exc}")
                if len(errors) >= 20:
                    errors.append(f"… (autres erreurs tronquées, total {len(rows)} lignes)")
                    break

        msg = f"Import {meta['label']} : {created} créé(s), {skipped} ignoré(s) (doublons)"
        if errors:
            msg += f", {len(errors)} erreur(s)"
            for e in errors[:5]:
                dj_messages.warning(request, e)
        dj_messages.success(request, msg)
        return redirect(meta["back_url"])

    def _create_employee(self, row, tenant):
        from employees.models import Employee
        matricule = (row.get("matricule") or "").strip()
        if not matricule:
            return False
        if Employee.objects.filter(tenant=tenant, matricule=matricule).exists():
            return False
        Employee.objects.create(
            tenant=tenant,
            matricule=matricule,
            first_name=(row.get("first_name") or "").strip(),
            last_name=(row.get("last_name") or "").strip(),
            email=(row.get("email") or "").strip(),
            phone=(row.get("phone") or "").strip(),
            contract_type=(row.get("contract_type") or "cdi").strip().lower(),
            status=(row.get("status") or "active").strip().lower(),
            work_location=(row.get("work_location") or "office").strip().lower(),
        )
        return True

    def _create_worker(self, row, tenant):
        from ouvriers.models import Subcontractor, Trade, Worker
        matricule = (row.get("matricule") or "").strip()
        if not matricule:
            return False
        if Worker.objects.filter(tenant=tenant, matricule=matricule).exists():
            return False
        trade_code = (row.get("trade_code") or "").strip().lower()
        trade = Trade.objects.filter(code=trade_code).first() if trade_code else None
        sub_code = (row.get("subcontractor_code") or "").strip().lower()
        sub = Subcontractor.objects.filter(
            tenant=tenant, code=sub_code).first() if sub_code else None
        Worker.objects.create(
            tenant=tenant, matricule=matricule,
            first_name=(row.get("first_name") or "").strip(),
            last_name=(row.get("last_name") or "").strip(),
            phone=(row.get("phone") or "").strip(),
            trade=trade, subcontractor=sub,
            status=(row.get("status") or "active").strip().lower(),
        )
        return True


class MyProfileRedirectView(View):
    """GET /me/<verb>/ → redirige vers la vue user_detail/update/password du user courant."""

    def get(self, request, verb="detail"):
        if not request.user.is_authenticated:
            return redirect("/auth/login/?next=/me/")
        pk = request.user.pk
        if verb == "detail":
            return redirect("admin-user-detail", pk=pk)
        if verb == "update":
            return redirect("admin-user-update", pk=pk)
        if verb == "password":
            return redirect("admin-user-password", pk=pk)
        return redirect("admin-user-detail", pk=pk)


class DataExportGenerateView(View):
    """POST /data-exports/<pk>/generate/ → produit le ZIP RGPD."""

    def post(self, request, pk):
        from audit.models import DataExportRequest
        from audit.services import generate_export_zip
        req = get_object_or_404(DataExportRequest, pk=pk)
        ok = generate_export_zip(req)
        if ok:
            try:
                from notifications.services import notify_export_ready
                notify_export_ready(req)
            except Exception:
                logger.warning("Notification export RGPD non envoyée (req=%s)", req.pk, exc_info=True)
            dj_messages.success(request, f"Export RGPD généré ({req.file.name}).")
        else:
            dj_messages.error(request, "Échec de la génération — vérifiez les logs.")
        return redirect("admin-audit")


class FraudAlertActionView(View):
    """POST /antifraud-alerts/<pk>/action/<verb>/ → transitions de statut."""

    VALID_VERBS = ("acknowledge", "dismiss", "escalate", "resolve")

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect("admin-login")
        from accounts.rbac import user_has_permission
        if not user_has_permission(request.user, "antifraud.acknowledge_alert"):
            dj_messages.error(request, "Permission refusée.")
            return redirect("admin-antifraud")
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, pk, verb):
        from django.utils import timezone

        from antifraud.models import FraudAlert, FraudInvestigation
        if verb not in self.VALID_VERBS:
            return redirect("admin-antifraud")
        alert = get_object_or_404(FraudAlert, pk=pk)
        now = timezone.now()

        if verb == "acknowledge" and alert.status == "open":
            alert.status = "acknowledged"
            alert.assigned_to = request.user if request.user.is_authenticated else None
            alert.save(update_fields=["status", "assigned_to"])
            dj_messages.success(request, "Alerte acquittée.")

        elif verb == "dismiss" and alert.status in ("open", "acknowledged"):
            alert.status = "dismissed"
            alert.resolution_comment = request.POST.get("comment", "Faux positif")
            alert.resolved_at = now
            alert.resolved_by = request.user if request.user.is_authenticated else None
            alert.save(update_fields=["status", "resolution_comment",
                                       "resolved_at", "resolved_by"])
            dj_messages.info(request, "Alerte marquée comme faux-positif.")

        elif verb == "escalate":
            # Crée une FraudInvestigation pour cette alerte
            FraudInvestigation.objects.create(
                tenant=alert.tenant,
                started_at=now,
                started_by=request.user if request.user.is_authenticated else None,
            ) if hasattr(FraudInvestigation, "started_by") else FraudInvestigation.objects.create(
                tenant=alert.tenant,
            )
            alert.status = "escalated"
            alert.save(update_fields=["status"])
            dj_messages.warning(request, "Alerte escaladée → enquête créée.")

        elif verb == "resolve":
            alert.status = "confirmed"
            alert.resolution_comment = request.POST.get("comment", "")
            alert.resolved_at = now
            alert.resolved_by = request.user if request.user.is_authenticated else None
            alert.save(update_fields=["status", "resolution_comment",
                                       "resolved_at", "resolved_by"])
            dj_messages.success(request, "Alerte confirmée et résolue.")

        return redirect("admin-antifraud")


# ===========================================================================
# AccessRule — vue liste dédiée (au-delà du CRUD)
# ===========================================================================
class AccessRulesView(BaseAdminView):
    template_name = "administration/access_rules.html"
    active_nav = "access"
    page_title = "Règles d'accès"
    page_subtitle = ("Configurez les règles appliquées par le moteur de décision : "
                      "horaires, zones réservées, certifications requises, etc.")
    breadcrumb = "Règles d'accès"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        try:
            from django.db.models import Count

            from access_control.models import AccessRule
            qs = (AccessRule.objects.select_related("site")
                  .annotate(events_count=Count("decisions"))
                  .order_by("-is_active", "site__name", "type"))
            page_obj, page_qs = paginate(qs, self.request, per_page=25)
            ctx["page_obj"] = page_obj
            ctx["rules"] = page_qs
            ctx["total"] = AccessRule.objects.count()
            ctx["active_count"] = AccessRule.objects.filter(is_active=True).count()
        except Exception:
            ctx["rules"] = []
            ctx["page_obj"] = None
            ctx["total"] = 0
            ctx["active_count"] = 0
        return ctx


# ─────────────────────────────────────────────────────────────────────
# Point 5 — Digest exécutif IA
# ─────────────────────────────────────────────────────────────────────
class ExecutiveDigestListView(BaseAdminView):
    """Liste des digests exécutifs générés (hebdo/mensuel/trim)."""
    template_name = "administration/digests.html"
    active_nav = "reports"
    page_title = "Digest exécutif IA"
    page_subtitle = ("Synthèses générées automatiquement chaque lundi par "
                      "l'assistant IA : KPI, alertes, anomalies et "
                      "recommandations.")
    breadcrumb = "Reporting · Digest IA"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        try:
            from reports.models import ExecutiveDigest
            qs = (ExecutiveDigest.objects.select_related("tenant")
                  .order_by("-period_start", "-created_at"))
            period = self.request.GET.get("period", "")
            if period in ("weekly", "monthly", "quarterly"):
                qs = qs.filter(period=period)
            status = self.request.GET.get("status", "")
            if status in ("queued", "generating", "ready", "failed"):
                qs = qs.filter(status=status)
            page_obj, page_qs = paginate(qs, self.request, per_page=20)
            ctx["page_obj"] = page_obj
            ctx["digests"] = page_qs

            ctx["counts"] = {
                "total": ExecutiveDigest.objects.count(),
                "ready": ExecutiveDigest.objects.filter(status="ready").count(),
                "failed": ExecutiveDigest.objects.filter(status="failed").count(),
                "weekly": ExecutiveDigest.objects.filter(period="weekly").count(),
                "monthly": ExecutiveDigest.objects.filter(period="monthly").count(),
            }
            ctx["filter_period"] = period
            ctx["filter_status"] = status
            ctx["last_ready"] = (ExecutiveDigest.objects
                                  .filter(status="ready")
                                  .order_by("-period_start").first())
        except Exception:
            logger.exception("ExecutiveDigestListView échec")
            ctx["digests"] = []
            ctx["page_obj"] = None
            ctx["counts"] = {"total": 0, "ready": 0, "failed": 0,
                              "weekly": 0, "monthly": 0}
            ctx["last_ready"] = None
        return ctx


class ExecutiveDigestDetailView(BaseAdminView):
    """Détail d'un digest exécutif : résumé, KPI, top alertes, recommandations."""
    template_name = "administration/digest_detail.html"
    active_nav = "reports"
    page_title = "Digest exécutif"
    breadcrumb = "Reporting · Digest"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        try:
            from reports.models import ExecutiveDigest
            digest = get_object_or_404(ExecutiveDigest, pk=kwargs["pk"])
            ctx["digest"] = digest
            ctx["page_title"] = digest.title or "Digest exécutif"
            ctx["breadcrumb"] = f"Reporting · {digest.get_period_display()} " \
                                  f"{digest.period_start}"
            # Pour l'affichage de l'email rendu (preview)
            try:
                from reports.services_digest import _render_digest_html
                _, ctx["email_html"] = _render_digest_html(digest)
            except Exception:
                ctx["email_html"] = ""
        except Exception:
            logger.exception("ExecutiveDigestDetailView échec")
            raise Http404("Digest introuvable")
        return ctx


class ExecutiveDigestActionView(LoginRequiredMixin, View):
    """Actions sur un digest : regenerate, send, generate_now."""

    def post(self, request, *args, **kwargs):
        from reports.models import ExecutiveDigest
        action_name = kwargs.get("verb", "")
        pk = kwargs.get("pk")

        if action_name == "generate":
            # Trigger global : on-demand pour le tenant courant
            from core.services import get_kaydan_tenant
            try:
                tenant = get_kaydan_tenant()
            except Exception:
                tenant = None
            if not tenant:
                dj_messages.error(request, "Tenant KAYDAN introuvable.")
                return redirect("admin-digests")
            period = request.POST.get("period", "weekly")
            try:
                from reports.tasks import generate_digest_for_tenant
                # Synchrone si CELERY_TASK_ALWAYS_EAGER, async sinon
                generate_digest_for_tenant.delay(
                    int(tenant.id), period=period, send_email=False)
                dj_messages.success(request,
                    f"Génération du digest {period} déclenchée.")
            except Exception as exc:
                dj_messages.error(request, f"Échec : {exc}")
            return redirect("admin-digests")

        if not pk:
            return redirect("admin-digests")

        digest = get_object_or_404(ExecutiveDigest, pk=pk)

        if action_name == "regenerate":
            try:
                from reports.tasks import regenerate_digest
                regenerate_digest.delay(int(digest.id), send_email=False)
                dj_messages.success(request,
                    "Régénération du digest mise en file.")
            except Exception as exc:
                dj_messages.error(request, f"Échec : {exc}")

        elif action_name == "send":
            try:
                from reports.services_digest import send_digest_email
                n = send_digest_email(digest)
                if n:
                    dj_messages.success(request,
                        f"Digest envoyé à {n} destinataire(s).")
                else:
                    dj_messages.warning(request,
                        "Aucun destinataire ou envoi échoué.")
            except Exception as exc:
                dj_messages.error(request, f"Échec envoi : {exc}")

        return redirect("admin-digest-detail", pk=digest.id)
